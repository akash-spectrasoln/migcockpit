from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.db import connection, transaction
from django.http import HttpResponse
from .models import SourceDB, SourceForm, Customer, Country, User   
from .serializers import SqlConnectionSerializer, SourceDbSerializer, SourceFormSerializer, CountrySerializer, SourceConnectionSerializer, DestinationConnectionSerializer, FileUploadSerializer, CustomerSerializer, UserSerializer
from .models import ValidationRules
from .serializers import ValidationRulesSerializer
from fetch_sqlserver.fetch_sqldata import extract_data
from encryption.encryption import encrypt_field, decrypt_field
from django.utils import timezone
from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.serializers import TokenRefreshSerializer
import hashlib
import json
import psycopg2
import pyodbc
from django.conf import settings
import pandas as pd
import io
import pycountry
import pgeocode
import zipcodes
import country_converter as cc
import pint
from api.authentications import JWTCookieAuthentication


def generate_encryption_key(cust_id, created_on):
    """
    Generate encryption key from cust_id + created_on date.
    Returns a hash that can be used as an integer for the encryption function.
    """
    # Convert created_on to string and combine with cust_id
    key_string = f"{cust_id}{created_on.strftime('%Y%m%d%H%M%S')}"
    
    # Generate a hash and convert to integer
    hash_object = hashlib.sha256(key_string.encode())
    # Take first 8 bytes and convert to integer (to fit within reasonable range)
    return int(hash_object.hexdigest()[:8], 16)

def create_connection_config(validated_data):
    """
    Create a clean JSON configuration from validated data, excluding null/empty values.
    """
    connection_config = {}
    
    # Define the fields to include in the config
    config_fields = ['hostname', 'user', 'password', 'schema', 'port']
    
    for field in config_fields:
        value = validated_data.get(field)
        # Only add non-null, non-empty values
        if value is not None and value != '':
            if field == 'port':
                connection_config[field] = value
            elif value.strip():  # For string fields, check if not empty after stripping
                connection_config[field] = value
    
    return connection_config

def test_database_connection(hostname, port, user, password, schema=None):
    """
    Test database connection with provided credentials.
    Supports both PostgreSQL and SQL Server connections.
    Returns (success, error_message)
    """
    try:
        # Validate required parameters
        if not hostname or not user or not password:
            return False, "Missing required connection parameters (hostname, user, password)"
        
        if port is None or port <= 0:
            return False, "Invalid port number"
        
        # Determine database type based on port
        port = int(port)
        if port == 1433:
            # SQL Server connection
            return test_sqlserver_connection(hostname, port, user, password, schema)
        elif port == 5432:
            # PostgreSQL connection
            return test_postgresql_connection(hostname, port, user, password, schema)
        else:
            # Try both PostgreSQL and SQL Server
            # First try SQL Server
            success, error = test_sqlserver_connection(hostname, port, user, password, schema)
            if success:
                return True, None
            
            # Then try PostgreSQL
            success, error = test_postgresql_connection(hostname, port, user, password, schema)
            if success:
                return True, None
            else:
                return False, f"Connection failed for both SQL Server and PostgreSQL. Last error: {error}"
        
    except ValueError as e:
        return False, f"Invalid parameter: {str(e)}"
    except Exception as e:
        return False, f"Unexpected error: {str(e)}"


def test_sqlserver_connection(hostname, port, user, password, schema=None):
    """
    Test SQL Server database connection.
    """
    try:
        # SQL Server connection string
        driver = 'ODBC Driver 17 for SQL Server'
        database = schema.strip() if schema and schema.strip() else 'master'
        
        conn_str = (
            f'Driver={driver};'
            f'Server={hostname.strip()},{port};'
            f'Database={database};'
            f'UID={user.strip()};'
            f'PWD={password};'
            f'Connection Timeout=10;'
        )
        
        conn = pyodbc.connect(conn_str)
        conn.close()
        return True, None
        
    except pyodbc.Error as e:
        error_msg = str(e)
        if "login failed" in error_msg.lower():
            return False, "Authentication failed: Invalid username or password"
        elif "could not connect" in error_msg.lower():
            return False, f"Connection failed: Unable to connect to {hostname}:{port}"
        elif "timeout" in error_msg.lower():
            return False, "Connection timeout: Server did not respond within 10 seconds"
        else:
            return False, f"SQL Server connection error: {error_msg}"
    except Exception as e:
        return False, f"SQL Server unexpected error: {str(e)}"


def test_postgresql_connection(hostname, port, user, password, schema=None):
    """
    Test PostgreSQL database connection.
    """
    try:
        # Build connection parameters
        conn_params = {
            'host': hostname.strip(),
            'port': port,
            'user': user.strip(),
            'password': password,
        }
        
        # Add database name if schema is provided, otherwise use 'postgres'
        if schema and schema.strip():
            conn_params['database'] = schema.strip()
        else:
            conn_params['database'] = 'postgres'
        
        # Test the connection with timeout
        conn = psycopg2.connect(
            **conn_params,
            connect_timeout=10  # 10 second timeout
        )
        conn.close()
        return True, None
        
    except psycopg2.OperationalError as e:
        error_msg = str(e)
        if "authentication failed" in error_msg.lower():
            return False, "Authentication failed: Invalid username or password"
        elif "could not connect to server" in error_msg.lower():
            return False, f"Connection failed: Unable to connect to {hostname}:{port}"
        elif "timeout expired" in error_msg.lower():
            return False, "Connection timeout: Server did not respond within 10 seconds"
        else:
            return False, f"PostgreSQL connection error: {error_msg}"
    except psycopg2.InterfaceError as e:
        return False, f"PostgreSQL interface error: {str(e)}"
    except Exception as e:
        return False, f"PostgreSQL unexpected error: {str(e)}"


def convert_user_date_format_to_strftime(user_date_format):
    """
    Convert user-friendly date format (e.g., 'MM-DD-YYYY') to Python strftime format (e.g., '%m-%d-%Y').
    
    Args:
        user_date_format: String format like 'MM-DD-YYYY', 'DD-MM-YYYY', etc.
    
    Returns:
        Python strftime format string
    """
    # Mapping of user-friendly format to strftime format
    format_mapping = {
        'YYYY': '%Y',
        'YY': '%y',
        'MM': '%m',
        'DD': '%d',
        'HH': '%H',
        'mm': '%M',
        'SS': '%S'
    }
    
    strftime_format = user_date_format
    for user_fmt, strftime_fmt in format_mapping.items():
        strftime_format = strftime_format.replace(user_fmt, strftime_fmt)
    
    return strftime_format


def format_date_columns(data, columns, date_format):
    """
    Format date and timestamp columns in the data based on column datatypes.
    
    Args:
        data: List of dictionaries containing row data
        columns: List of tuples containing (column_name, data_type, is_nullable)
        date_format: String format for output dates in strftime format
    
    Returns:
        Formatted data with dates properly formatted
    """
    if not data:
        return data
    
    # Identify date/timestamp columns based on datatype
    date_columns = []
    time_columns = []  # Track columns that should include time
    
    for col in columns:
        col_name, col_type = col[0], col[1].lower()
        # Check for date, timestamp, or datetime datatypes (including PostgreSQL variations)
        if any(dtype in col_type for dtype in ['timestamp', 'datetime']):
            # Timestamp columns may include time information
            date_columns.append(col_name)
            time_columns.append(col_name)
        elif 'date' in col_type and 'time' not in col_type:
            # Pure date columns (not timestamp)
            date_columns.append(col_name)
    
    # If no date columns found, return data as is
    if not date_columns:
        return data
    
    # Format date columns in each row
    for row in data:
        for col_name in date_columns:
            if col_name in row and row[col_name] is not None:
                value = row[col_name]
                try:
                    # Determine format to use based on whether this column should include time
                    format_to_use = date_format
                    
                    # If value is a string (ISO format from database), parse it first
                    if isinstance(value, str):
                        # Try parsing ISO format or other common formats
                        import dateutil.parser
                        parsed_date = dateutil.parser.parse(value)
                        
                        # For timestamp columns, check if time component exists and is not midnight
                        if col_name in time_columns:
                            # Check if the time component is not 00:00:00
                            if parsed_date.hour != 0 or parsed_date.minute != 0 or parsed_date.second != 0:
                                # Append time format to date format
                                format_to_use = f"{date_format} %H:%M:%S"
                        
                        row[col_name] = parsed_date.strftime(format_to_use)
                    # If value has strftime method (datetime object), format directly
                    elif hasattr(value, 'strftime'):
                        # For timestamp columns, check if time component exists and is not midnight
                        if col_name in time_columns:
                            # Check if the time component is not 00:00:00
                            if value.hour != 0 or value.minute != 0 or value.second != 0:
                                # Append time format to date format
                                format_to_use = f"{date_format} %H:%M:%S"
                        
                        row[col_name] = value.strftime(format_to_use)
                except Exception:
                    # If formatting fails, convert to ISO format string for JSON serialization
                    try:
                        if hasattr(value, 'isoformat'):
                            row[col_name] = value.isoformat()
                    except:
                        # Last resort: keep original value
                        continue
    
    return data


def decrypt_source_data(encrypted_data, cust_id, created_on):
    """
    Decrypt source data using the same key generation logic.
    """
    if not encrypted_data:
        return None
    
    try:
        # Parse the JSON data
        data_list = json.loads(encrypted_data) if isinstance(encrypted_data, str) else encrypted_data
        
        # Generate the same encryption key
        encryption_key = generate_encryption_key(cust_id, created_on)
        
        # Decrypt the field
        return decrypt_field(
            data_list[0],  # encrypted_data
            encryption_key,  # cmp_id
            data_list[1],   # nonce
            data_list[2],   # tag
            data_list[3],   # salt
            data_list[4],   # original_type
            data_list[5]    # iterations
        )
    except Exception as e:
        print(f"Error decrypting data: {e}")
        return None

class SqlConnectionView(APIView):
    def post(self, request):
        serializer = SqlConnectionSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            success, message = extract_data(
                data['sql_hostname'],
                data['sql_database'],
                data['sql_username'],
                data['sql_password'],
                1433
            )
            if success:
                return Response({"message": "success"}, status=status.HTTP_201_CREATED)
            return Response({"message": "failed", "error": message}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
class SourcesListView(APIView):
    def get(self, request):
        sources = SourceDB.objects.all()
        return Response(SourceDbSerializer(sources, many=True).data, status=status.HTTP_200_OK)


class SourceFieldsView(APIView):
    def get(self, request, source_id: int):
        try:
            source = SourceDB.objects.get(id=source_id)
            fields = SourceForm.objects.filter(src_db=source)
        except SourceDB.DoesNotExist:
            return Response({"detail": "Source not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(SourceFormSerializer(fields, many=True).data, status=status.HTTP_200_OK)


class CountryListView(APIView):
    """API view for listing all countries."""
    
    def get(self, request):
        """Get list of all countries."""
        try:
            countries = Country.objects.all()
            serializer = CountrySerializer(countries, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

            
        except Exception as e:
            return Response(
                {"error": f"Error retrieving countries: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SourceConnectionCreateView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = SourceConnectionSerializer(data=request.data)
        if serializer.is_valid():
            try:
                #Get the customer object
                user = request.user
                customer = user.cust_id

                # Test database connection before saving
                hostname = serializer.validated_data.get("hostname")
                port = serializer.validated_data.get("port")
                user = serializer.validated_data.get("user")
                password = serializer.validated_data.get("password")
                schema = serializer.validated_data.get("schema")
                
                # # Test connection if any connection details are provided
                # if any([hostname, port, user, password]):
                #     connection_success, error_message = test_database_connection(
                #         hostname, port, user, password, schema
                #     )
                    
                #     if not connection_success:
                #         return Response(
                #             {
                #                 "error": "Database connection test failed",
                #                 "details": error_message,
                #                 "connection_params": {
                #                     "hostname": hostname,
                #                     "port": port,
                #                     "user": user,
                #                     "schema": schema
                #                 }
                #             }, 
                #             status=status.HTTP_400_BAD_REQUEST
                #         )

                # Create JSON structure with connection details (excluding id and name)
                # Only include non-null values in the JSON config
                connection_config = create_connection_config(serializer.validated_data)

                # Connect to the customer's database and insert into GENERAL.source
                import psycopg2
                import json
                from django.conf import settings

                conn = psycopg2.connect(
                    host=settings.DATABASES['default']['HOST'],
                    port=settings.DATABASES['default']['PORT'],
                    user=settings.DATABASES['default']['USER'],
                    password=settings.DATABASES['default']['PASSWORD'],
                    database=customer.cust_db
                )
                conn.autocommit = True
                cursor = conn.cursor()

                # Get the current timestamp from the database to ensure consistency
                cursor.execute("SELECT NOW()")
                db_timestamp = cursor.fetchone()[0]
                
                # Generate encryption key using the database timestamp
                encryption_key = generate_encryption_key(customer.cust_id, db_timestamp)
                
                # Encrypt the entire JSON configuration using the database timestamp
                encrypted_config = encrypt_field(connection_config, encryption_key)

                insert_sql = '''
                    INSERT INTO "GENERAL".source (source_name, source_config, created_on, modified_on, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                '''
                cursor.execute(
                    insert_sql,
                    (
                        serializer.validated_data.get("source_name"),
                        json.dumps(encrypted_config),
                        db_timestamp,
                        db_timestamp,
                        True  # is_active defaults to True
                    )
                )

                # After encrypting and storing, you can fetch and decrypt the data as follows:
                # Fetch the just-inserted source record for demonstration (e.g., by source_name)
                cursor.execute(
                    '''
                    SELECT source_name, source_config, created_on, modified_on, is_active
                    FROM "GENERAL".source
                    WHERE source_name = %s
                    ORDER BY id DESC LIMIT 1
                    ''',
                    (serializer.validated_data.get("source_name"),)
                )
                fetched_row = cursor.fetchone()
                if fetched_row:
                    fetched_source = dict(
                        source_name=fetched_row[0],
                        source_config=fetched_row[1],
                        created_on=fetched_row[2],
                        modified_on=fetched_row[3],
                        is_active=fetched_row[4],
                    )

                    # Decrypt the entire JSON configuration using source's created_on time
                    decrypted_config = decrypt_source_data(fetched_source["source_config"], customer.cust_id, fetched_source["created_on"]) if fetched_source["source_config"] else None

                    # Configuration successfully encrypted and stored

                cursor.close()
                conn.close()
                
                return Response({"message": "Source connection added successfully"}, status=status.HTTP_201_CREATED)
                
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            except Exception as e:
                return Response(
                    {"error": f"Failed to store source details in customer database: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CustomerSourcesView(APIView):

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request,):
        try:
            # Get the customer object
            user = request.user
            customer = user.cust_id
            
            # Connect to the customer's database and fetch sources
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()

            # Fetch sources from the customer's database
            cursor.execute('''
                SELECT id, source_name, source_config, created_on, modified_on, is_active
                FROM "GENERAL".source
                ORDER BY created_on DESC
            ''')
            
            sources = []
            for row in cursor.fetchall():
                source_id, source_name, source_config, created_on, modified_on, is_active = row
                
                # Decrypt the source configuration
                decrypted_config = {}
                if source_config:
                    try:
                        decrypted_config = decrypt_source_data(source_config, customer.cust_id, created_on)
                        
                    except Exception as e:
                        print(f"Error decrypting source config: {e}")
                        decrypted_config = {}
                
                source_data = {
                    'source_id': source_id,
                    'source_name': source_name,
                    'hostname': decrypted_config.get('hostname') if decrypted_config else None,
                    'port': decrypted_config.get('port') if decrypted_config else None,
                    'user': decrypted_config.get('user') if decrypted_config else None,
                    'password': decrypted_config.get('password') if decrypted_config else None,
                    'schema': decrypted_config.get('schema') if decrypted_config else None,
                    'created_on': created_on.isoformat() if created_on else None,
                    'modified_on': modified_on.isoformat() if modified_on else None,
                    'is_active': is_active
                }
                sources.append(source_data)

            cursor.close()
            conn.close()
            
            return Response({
                'customer_id': customer.cust_id,
                'customer_name': customer.name,
                'sources': sources
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to fetch sources: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SourceEditView(APIView):
    """API view for editing existing source connections."""
    
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, source_id):
        """Get source data for editing."""
        try:
            # Get the customer object
            user = request.user
            customer = user.cust_id
            
            # Connect to the customer's database
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()

            # Fetch the specific source
            cursor.execute('''
                SELECT source_name, source_config, created_on, modified_on, is_active
                FROM "GENERAL".source
                WHERE id = %s
            ''', (source_id,))
            
            source_row = cursor.fetchone()
            if not source_row:
                return Response(
                    {"error": "Source not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            source_name_db, source_config, created_on, modified_on, is_active = source_row
            
            # Decrypt the source configuration
            decrypted_config = {}
            if source_config:
                try:
                    decrypted_config = decrypt_source_data(source_config, customer.cust_id, created_on)
                except Exception as e:
                    print(f"Error decrypting source config: {e}")
                    decrypted_config = {}
            
            source_data = {
                'source_name': source_name_db,
                'hostname': decrypted_config.get('hostname') if decrypted_config else '',
                'port': decrypted_config.get('port') if decrypted_config else '',
                'user': decrypted_config.get('user') if decrypted_config else '',
                'password': decrypted_config.get('password') if decrypted_config else '',
                'schema': decrypted_config.get('schema') if decrypted_config else '',
                'created_on': created_on.isoformat() if created_on else None,
                'modified_on': modified_on.isoformat() if modified_on else None,
                'is_active': is_active
            }

            cursor.close()
            conn.close()
            
            return Response({
                'customer_id': customer.cust_id,
                'customer_name': customer.name,
                'source': source_data
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to fetch source data: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def put(self, request, source_id):
        """Update an existing source connection."""
        try:
            # Get the customer object
            user = request.user
            customer = user.cust_id
            
            # Connect to the customer's database
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()

            # Check if source exists
            cursor.execute('''
                SELECT source_name, source_config, created_on, is_active
                FROM "GENERAL".source
                WHERE id = %s
            ''', (source_id,))
            
            existing_source = cursor.fetchone()
            if not existing_source:
                return Response(
                    {"error": "Source not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            # Get the original created_on timestamp to maintain encryption consistency
            original_source_name, original_source_config, original_created_on, original_is_active = existing_source


            # Create new connection configuration
            connection_config = create_connection_config(request.data)

            # Get current timestamp for modification
            cursor.execute("SELECT NOW()")
            current_timestamp = cursor.fetchone()[0]

            # Use the original created_on timestamp for encryption key generation
            # This ensures we can still decrypt the data
            encryption_key = generate_encryption_key(customer.cust_id, original_created_on)
            
            # Encrypt the new configuration using the same key
            encrypted_config = encrypt_field(connection_config, encryption_key)

            # Update the source record
            update_sql = '''
                UPDATE "GENERAL".source 
                SET source_name = %s, 
                    source_config = %s, 
                    modified_on = %s, 
                    is_active = %s
                WHERE id = %s
            '''
            
            cursor.execute(
                update_sql,
                (
                    request.data.get('source_name'),
                    json.dumps(encrypted_config),
                    current_timestamp,
                    request.data.get('is_active', True),
                    source_id  # Original source name for WHERE clause
                )
            )

            # Verify the update by fetching the updated record
            cursor.execute('''
                SELECT source_name, source_config, created_on, modified_on, is_active
                FROM "GENERAL".source
                WHERE id = %s
            ''', (source_id,))
            
            updated_source = cursor.fetchone()
            if updated_source:
                updated_source_name, updated_source_config, created_on, modified_on, is_active = updated_source
                
                # Configuration successfully updated

            cursor.close()
            conn.close()
            
            return Response({
                "message": "Source connection updated successfully",
                "modified_on": current_timestamp.isoformat()
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to update source connection: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SourceDeleteView(APIView):
    """API view for deleting existing source connections."""

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, source_id):
        """Delete an existing source connection."""
        try:
            # Get the customer object
            user = request.user
            customer = user.cust_id
            
            # Connect to the customer's database
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()

            # Check if source exists
            cursor.execute('''
                SELECT source_name, created_on
                FROM "GENERAL".source
                WHERE id = %s
            ''', (source_id,))
            
            existing_source = cursor.fetchone()
            if not existing_source:
                return Response(
                    {"error": "Source not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            # Delete the source record
            delete_sql = '''
                DELETE FROM "GENERAL".source 
                WHERE id = %s
            '''
            
            cursor.execute(delete_sql, (source_id,))
            
            # Verify the deletion
            cursor.execute('''
                SELECT COUNT(*) FROM "GENERAL".source 
                WHERE id = %s
            ''', (source_id,))
            
            remaining_count = cursor.fetchone()[0]
            if remaining_count > 0:
                return Response(
                    {"error": "Failed to delete source"}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            cursor.close()
            conn.close()
            
            return Response({
                "message": "Source connection deleted successfully",
                "source_id": source_id
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to delete source connection: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DestinationConnectionCreateView(APIView):
    """API view for creating destination connections."""
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Create a new destination connection."""
        serializer = DestinationConnectionSerializer(data=request.data)
        if serializer.is_valid():
            try:
                # Get the customer object
                user = request.user
                customer = user.cust_id

                # Create destination configuration
                destination_config = {
                    'hostname': serializer.validated_data.get("hostname"),
                    'instance_number': serializer.validated_data.get("instance_number"),
                    'mode': serializer.validated_data.get("mode"),
                    'destination_schema_name': serializer.validated_data.get("destination_schema_name"),
                    's4_schema_name': serializer.validated_data.get("s4_schema_name"),
                }
                
                # Add database type and name based on mode
                if serializer.validated_data.get("mode") == 'multiple_containers':
                    destination_config['database_type'] = serializer.validated_data.get("database_type")
                    
                    # Add the appropriate database name based on database_type
                    if serializer.validated_data.get("database_type") == 'tenant_database':
                        destination_config['tenant_db_name'] = serializer.validated_data.get("tenant_db_name")
                    elif serializer.validated_data.get("database_type") == 'system_database':
                        destination_config['system_db_name'] = serializer.validated_data.get("system_db_name")

                # Connect to the customer's database and insert into GENERAL.destination
                import psycopg2
                import json
                from django.conf import settings

                conn = psycopg2.connect(
                    host=settings.DATABASES['default']['HOST'],
                    port=settings.DATABASES['default']['PORT'],
                    user=settings.DATABASES['default']['USER'],
                    password=settings.DATABASES['default']['PASSWORD'],
                    database=customer.cust_db
                )
                conn.autocommit = True
                cursor = conn.cursor()

                # Get the current timestamp from the database to ensure consistency
                cursor.execute("SELECT NOW()")
                db_timestamp = cursor.fetchone()[0]
                
                # Generate encryption key using the database timestamp
                encryption_key = generate_encryption_key(customer.cust_id, db_timestamp)
                
                # Encrypt the entire JSON configuration using the database timestamp
                encrypted_config = encrypt_field(destination_config, encryption_key)

                insert_sql = '''
                    INSERT INTO "GENERAL".destination (destination_name, destination_config, created_on, modified_on, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                '''
                cursor.execute(
                    insert_sql,
                    (
                        serializer.validated_data.get("destination_name"),
                        json.dumps(encrypted_config),
                        db_timestamp,
                        db_timestamp,
                        True  # is_active defaults to True
                    )
                )

                # Verify the insertion by fetching the just-inserted destination record
                cursor.execute(
                    '''
                    SELECT destination_name, destination_config, created_on, modified_on, is_active
                    FROM "GENERAL".destination
                    WHERE destination_name = %s
                    ORDER BY id DESC LIMIT 1
                    ''',
                    (serializer.validated_data.get("destination_name"),)
                )
                fetched_row = cursor.fetchone()
                if fetched_row:
                    fetched_destination = dict(
                        destination_name=fetched_row[0],
                        destination_config=fetched_row[1],
                        created_on=fetched_row[2],
                        modified_on=fetched_row[3],
                        is_active=fetched_row[4],
                    )

                    # Decrypt the entire JSON configuration using destination's created_on time
                    decrypted_config = decrypt_source_data(fetched_destination["destination_config"], customer.cust_id, fetched_destination["created_on"]) if fetched_destination["destination_config"] else None

                    # Configuration successfully encrypted and stored

                cursor.close()
                conn.close()
                
                return Response({
                    "message": "Destination connection added successfully",
                    "destination_name": serializer.validated_data.get("destination_name"),
                    "mode": serializer.validated_data.get("mode")
                }, status=status.HTTP_201_CREATED)
                
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            except Exception as e:
                return Response(
                    {"error": f"Failed to store destination details in customer database: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CustomerDestinationsView(APIView):
    """
    API view to fetch destinations for a specific customer.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get the customer
            user = request.user
            customer = user.cust_id
            
            # Connect to the customer's database
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()
            
            # Query destinations from the customer's database
            cursor.execute('''
                SELECT id, destination_name, destination_config, created_on, modified_on, is_active
                FROM "GENERAL".destination
                ORDER BY created_on DESC
            ''')
            
            destinations = []
            for row in cursor.fetchall():
                destination_id, destination_name, destination_config, created_on, modified_on, is_active = row
                
                # Decrypt the destination configuration
                decrypted_config = {}
                if destination_config:
                    try:
                        decrypted_config = decrypt_source_data(destination_config, customer.cust_id, created_on)
                    except Exception as e:
                        print(f"Error decrypting destination config: {e}")
                        decrypted_config = {}
                
                destination_data = {
                    'destination_id': destination_id,
                    'destination_name': destination_name,
                    'hostname': decrypted_config.get('hostname') if decrypted_config else None,
                    'port': decrypted_config.get('instance_number') if decrypted_config else None,  # Using instance_number as port
                    'user': None,  # Destinations don't have user field
                    'password': None,  # Destinations don't have password field
                    'schema': decrypted_config.get('destination_schema_name') if decrypted_config else None,
                    'mode': decrypted_config.get('mode') if decrypted_config else None,
                    'database_type': decrypted_config.get('database_type') if decrypted_config else None,
                    'tenant_db_name': decrypted_config.get('tenant_db_name') if decrypted_config else None,
                    'system_db_name': decrypted_config.get('system_db_name') if decrypted_config else None,
                    's4_schema_name': decrypted_config.get('s4_schema_name') if decrypted_config else None,
                    'created_on': created_on.isoformat() if created_on else None,
                    'modified_on': modified_on.isoformat() if modified_on else None,
                    'is_active': is_active
                }
                destinations.append(destination_data)

            cursor.close()
            conn.close()
            
            return Response({
                'customer_id': customer.cust_id,
                'customer_name': customer.name,
                'destinations': destinations
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {'error': 'Customer not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"Error fetching customer destinations: {e}")
            return Response(
                {'error': 'Failed to fetch destinations'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DestinationEditView(APIView):
    """API view for editing existing destination connections."""

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, destination_id):
        """Get destination data for editing."""
        try:
            # Get the customer object]
            user = request.user
            customer = user.cust_id
            
            # Connect to the customer's database
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()

            # Fetch the specific destination
            cursor.execute('''
                SELECT destination_name, destination_config, created_on, modified_on, is_active
                FROM "GENERAL".destination
                WHERE id = %s
            ''', (destination_id,))
            
            destination_row = cursor.fetchone()
            if not destination_row:
                return Response(
                    {"error": "Destination not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            destination_name_db, destination_config, created_on, modified_on, is_active = destination_row
            
            # Decrypt the destination configuration
            decrypted_config = {}
            if destination_config:
                try:
                    decrypted_config = decrypt_source_data(destination_config, customer.cust_id, created_on)
                except Exception as e:
                    print(f"Error decrypting destination config: {e}")
                    decrypted_config = {}
            
            destination_data = {
                'destination_name': destination_name_db,
                'hostname': decrypted_config.get('hostname') if decrypted_config else '',
                'instance_number': decrypted_config.get('instance_number') if decrypted_config else '',
                'mode': decrypted_config.get('mode') if decrypted_config else '',
                'database_type': decrypted_config.get('database_type') if decrypted_config else '',
                'tenant_db_name': decrypted_config.get('tenant_db_name') if decrypted_config else '',
                'system_db_name': decrypted_config.get('system_db_name') if decrypted_config else '',
                'destination_schema_name': decrypted_config.get('destination_schema_name') if decrypted_config else '',
                's4_schema_name': decrypted_config.get('s4_schema_name') if decrypted_config else '',
                'created_on': created_on.isoformat() if created_on else None,
                'modified_on': modified_on.isoformat() if modified_on else None,
                'is_active': is_active
            }

            cursor.close()
            conn.close()
            
            return Response({
                'customer_id': customer.cust_id,
                'customer_name': customer.name,
                'destination': destination_data
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to fetch destination data: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def put(self, request, destination_id):
        """Update an existing destination connection."""
        try:
            # Get the customer object
            user = request.user
            customer = user.cust_id
            
            # Connect to the customer's database
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()

            # Check if destination exists
            cursor.execute('''
                SELECT destination_name, destination_config, created_on, is_active
                FROM "GENERAL".destination
                WHERE id = %s
            ''', (destination_id,))
            
            existing_destination = cursor.fetchone()
            if not existing_destination:
                return Response(
                    {"error": "Destination not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            # Get the original created_on timestamp to maintain encryption consistency
            original_destination_name, original_destination_config, original_created_on, original_is_active = existing_destination

            # Create new destination configuration
            destination_config = {
                'hostname': request.data.get('hostname'),
                'instance_number': request.data.get('instance_number'),
                'mode': request.data.get('mode'),
                'destination_schema_name': request.data.get('destination_schema_name'),
                's4_schema_name': request.data.get('s4_schema_name'),
            }
            
            # Add database type and name based on mode
            if request.data.get('mode') == 'multiple_containers':
                destination_config['database_type'] = request.data.get('database_type')
                
                # Add the appropriate database name based on database_type
                if request.data.get('database_type') == 'tenant_database':
                    destination_config['tenant_db_name'] = request.data.get('tenant_db_name')
                elif request.data.get('database_type') == 'system_database':
                    destination_config['system_db_name'] = request.data.get('system_db_name')

            # Get current timestamp for modification
            cursor.execute("SELECT NOW()")
            current_timestamp = cursor.fetchone()[0]

            # Use the original created_on timestamp for encryption key generation
            # This ensures we can still decrypt the data
            encryption_key = generate_encryption_key(customer.cust_id, original_created_on)
            
            # Encrypt the new configuration using the same key
            encrypted_config = encrypt_field(destination_config, encryption_key)

            # Update the destination record
            update_sql = '''
                UPDATE "GENERAL".destination 
                SET destination_name = %s, 
                    destination_config = %s, 
                    modified_on = %s, 
                    is_active = %s
                WHERE id = %s
            '''
            
            cursor.execute(
                update_sql,
                (
                    request.data.get('destination_name'),
                    json.dumps(encrypted_config),
                    current_timestamp,
                    request.data.get('is_active', True),
                    destination_id  # Original destination id for WHERE clause
                )
            )

            # Verify the update by fetching the updated record
            cursor.execute('''
                SELECT destination_name, destination_config, created_on, modified_on, is_active
                FROM "GENERAL".destination
                WHERE id = %s
            ''', (destination_id,))
            
            updated_destination = cursor.fetchone()
            if updated_destination:
                destination_name, updated_destination_config, created_on, modified_on, is_active = updated_destination
                
                # Configuration successfully updated

            cursor.close()
            conn.close()
            
            return Response({
                "message": "Destination connection updated successfully",
                "modified_on": current_timestamp.isoformat()
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to update destination connection: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DestinationDeleteView(APIView):
    """API view for deleting existing destination connections."""
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, destination_id):
        """Delete an existing destination connection."""
        try:
            # Get the customer object
            customer = request.user.cust_id
            
            # Connect to the customer's database
            conn = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            conn.autocommit = True
            cursor = conn.cursor()

            # Check if destination exists
            cursor.execute('''
                SELECT destination_name, created_on
                FROM "GENERAL".destination
                WHERE id = %s
            ''', (destination_id,))
            
            existing_destination = cursor.fetchone()
            if not existing_destination:
                return Response(
                    {"error": "Destination not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            # Delete the destination record
            delete_sql = '''
                DELETE FROM "GENERAL".destination 
                WHERE id = %s
            '''
            
            cursor.execute(delete_sql, (destination_id,))
            
            # Verify the deletion
            cursor.execute('''
                SELECT COUNT(*) FROM "GENERAL".destination 
                WHERE id = %s
            ''', (destination_id,))
            
            remaining_count = cursor.fetchone()[0]
            if remaining_count > 0:
                return Response(
                    {"error": "Failed to delete destination"}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            cursor.close()
            conn.close()
            
            return Response({
                "message": "Destination connection deleted successfully",
            }, status=status.HTTP_200_OK)
            
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to delete destination connection: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class FileUploadPreviewView(APIView):
    """
    API view for uploading files and previewing table structure without creating the table.
    This allows users to review and modify column types before table creation.
"""

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value
        
        str_value = str(value).strip()
        
        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)
                
                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return f"{float_val:.10f}".rstrip('0').rstrip('.')
                else:
                    # Default: convert to int (for backward compatibility with phone numbers)
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value
        
        return str_value

    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()
        
        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)
        
        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')
        
        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)
        
        return df_cleaned
    
    def _detect_date_format(self, series):
        """
        Detect the consistent date format for a column by analyzing unambiguous dates.
        Returns the format string that should be used, or None if no format detected.
        """
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return None
        
        # Sample for performance
        sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values
        
        # Try each format and score based on successful conversions
        formats_to_try = [
            '%Y-%m-%d',                  # 2024-01-15
            '%m/%d/%Y',                  # 01/15/2024 (US format - month first)
            '%d/%m/%Y',                  # 15/01/2024 (European format - day first)
            '%m-%d-%Y',                  # 01-15-2024 (US format - month first)
            '%d-%m-%Y',                  # 15-01-2024 (European format - day first)
            '%Y/%m/%d',                  # 2024/01/15
            '%Y%m%d',                    # 20240115
            '%d.%m.%Y',                  # 15.01.2024
            '%Y-%m-%d %H:%M:%S',         # 2024-01-15 14:30:00
            '%m/%d/%Y %H:%M:%S',         # 01/15/2024 14:30:00
            '%d/%m/%Y %H:%M:%S',         # 15/01/2024 14:30:00
            '%m-%d-%Y %H:%M:%S',         # 01-15-2024 14:30:00
            '%d-%m-%Y %H:%M:%S',         # 15-01-2024 14:30:00
            '%Y-%m-%d %H:%M:%S.%f',      # 2024-01-15 14:30:00.123456
            '%m-%d-%Y %H:%M:%S.%f',      # 01-15-2024 14:30:00.123456
            '%d-%m-%Y %H:%M:%S.%f',      # 15-01-2024 14:30:00.123456
            '%m/%d/%Y %H:%M:%S.%f',      # 01/15/2024 14:30:00.123456
            '%d/%m/%Y %H:%M:%S.%f',      # 15/01/2024 14:30:00.123456
        ]
        
        best_format = None
        best_score = 0
        
        for fmt in formats_to_try:
            try:
                converted = pd.to_datetime(sample, format=fmt, errors='coerce')
                success_count = converted.notna().sum()
                
                # Calculate success rate
                if success_count > best_score:
                    best_score = success_count
                    best_format = fmt
            except (ValueError, TypeError):
                continue
        
        # Return format if at least 60% success rate
        if best_score / len(sample) > 0.6:
            return best_format
        
        return None
    
    def _is_datetime_column(self, series):
        """
        Efficiently detect if a pandas Series contains datetime values.
        Uses format detection to ensure consistent parsing.
        """
        # Skip empty or all-null columns
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return False
        
        # If already datetime type, return True
        if pd.api.types.is_datetime64_any_dtype(series):
            return True
        
        # Try to detect the format
        detected_format = self._detect_date_format(series)
        if detected_format:
            return True
        
        # Fallback: Try with infer_datetime_format
        try:
            sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values
            converted = pd.to_datetime(sample, errors='coerce', infer_datetime_format=True)
            success_rate = converted.notna().sum() / len(sample)
            return success_rate > 0.6
        except (ValueError, TypeError):
            return False
    
    def _detect_datetime_type(self, series):
        """
        Determine if datetime column should be DATE or TIMESTAMP.
        Returns 'DATE' if all times are midnight (00:00:00), otherwise 'TIMESTAMP'.
        Uses consistent format detection.
        """
        try:
            # First, try to detect the consistent format
            detected_format = self._detect_date_format(series)
            
            if detected_format:
                # Use the detected format for consistent parsing
                datetime_series = pd.to_datetime(series, format=detected_format, errors='coerce')
            else:
                # Fallback to infer_datetime_format
                datetime_series = pd.to_datetime(series, errors='coerce', infer_datetime_format=True)
            
            non_null_datetimes = datetime_series.dropna()
            
            if len(non_null_datetimes) == 0:
                return 'TIMESTAMP'  # Default to TIMESTAMP if no valid dates
            
            # Sample for performance (check first 100 rows)
            sample = non_null_datetimes.head(100) if len(non_null_datetimes) > 100 else non_null_datetimes
            
            # Check if all times are at midnight (00:00:00)
            # This indicates date-only data
            has_time_component = any(
                dt.hour != 0 or dt.minute != 0 or dt.second != 0 or dt.microsecond != 0
                for dt in sample
            )
            
            # Return DATE if no time component, TIMESTAMP if has time
            return 'TIMESTAMP' if has_time_component else 'DATE'
        except Exception:
            return 'TIMESTAMP'  # Default to TIMESTAMP on error

    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)

        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            file_extension = uploaded_file.name.lower().split('.')[-1]


            # Get user and customer information
            try:
                user = request.user
                customer = user.cust_id
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            try:
                uploaded_file.seek(0)

                # Read file into pandas DataFrame
                # Don't use dtype=str to allow pandas to infer types (especially dates)
                if file_extension == 'csv':
                    df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
                elif file_extension in ['xls', 'xlsx']:
                    df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
                else:
                    return Response(
                        {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Remove empty rows from DataFrame
                df = self._remove_empty_rows(df)

                # Clean DataFrame
                df_clean = df.where(pd.notnull(df), None)
                
                # Check for reserved field names
                reserved_fields = ['__id', 'is_active']
                found_reserved_fields = []
                for col in df_clean.columns:
                    col_lower = col.lower().strip()
                    if col_lower in reserved_fields or col_lower == 'id':
                        found_reserved_fields.append(col)
                
                if found_reserved_fields:
                    return Response(
                        {"error": f"The following field names are reserved and cannot be used: {', '.join(found_reserved_fields)}. Please rename these columns in your file and try again."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Fix scientific notation for all columns that might contain it
                for col in df_clean.columns:
                    # Check if any value in this column contains scientific notation
                    has_scientific_notation = False
                    for val in df_clean[col].dropna():
                        if isinstance(val, str) and 'E' in val.upper():
                            has_scientific_notation = True
                            break
                    
                    if has_scientific_notation:
                        # Convert scientific notation back to regular numbers
                        df_clean[col] = df_clean[col].apply(lambda x: self._fix_scientific_notation(x) if pd.notna(x) else x)
                
                # Define table name with filename__ format
                base_name = uploaded_file.name.rsplit('.', 1)[0]
                table_name = f"{base_name.lower()}__".replace(" ", "_").replace("-", "_")

                # Detect column types by checking all data in each column
                column_info = []
                datetime_columns = {}  # Store columns that are datetime and their detected format
                
                for col in df_clean.columns:
                    col_name = col.strip().replace(" ", "_").replace("-", "_").lower()
                    col_lower = col.lower()
                    # Get all non-null values for accurate type detection
                    all_values = df_clean[col].dropna()
                    # Still get sample values for display purposes
                    sample_values = all_values.head(10)
                    
                    # Enhanced type detection using ALL values in the column
                    # First, calculate max length for VARCHAR sizing
                    max_length = df_clean[col].astype(str).str.len().max()
                    varchar_length = max_length + 50  # Add exactly 50 to the biggest data length
                    
                    col_type = None  # Initialize as None to track if type was detected
                    
                    # Priority 1: Check for date/timestamp columns using pandas
                    if self._is_datetime_column(df_clean[col]):
                        col_type = self._detect_datetime_type(df_clean[col])
                        # Store the detected format for this column
                        detected_format = self._detect_date_format(df_clean[col])
                        datetime_columns[col] = detected_format
                    # Priority 2: Check for phone numbers
                    elif any(keyword in col_lower for keyword in ['phone', 'mobile', 'tel', 'contact']):
                        col_type = 'VARCHAR(20)'
                    # Priority 3: Check for values with + sign (likely phone numbers with country code)
                    elif any('+' in str(val) for val in all_values if pd.notna(val)):
                        col_type = f'VARCHAR({varchar_length})'
                    # Priority 4: Check for integers
                    elif all(str(val).isdigit() for val in all_values if pd.notna(val) and str(val) != ''):
                        # Check if numbers are within INTEGER range (-2,147,483,648 to 2,147,483,647)
                        max_int = 2147483647
                        min_int = -2147483648
                        try:
                            numeric_values = [int(val) for val in all_values if pd.notna(val) and str(val) != '']
                            if all(min_int <= val <= max_int for val in numeric_values):
                                col_type = 'INTEGER'
                            else:
                                col_type = 'BIGINT'  # Use BIGINT for large integers
                        except (ValueError, OverflowError):
                            col_type = 'BIGINT'  # Use BIGINT if conversion fails
                    # Priority 5: Check for decimal numbers
                    elif all(str(val).replace('.', '').isdigit() and str(val).count('.') <= 1 for val in all_values if pd.notna(val) and str(val) != ''):
                        try:
                            decimal_values = [float(val) for val in all_values if pd.notna(val) and str(val) != '']
                            # Check if all values are whole numbers (no decimal part)
                            if all(val.is_integer() for val in decimal_values):
                                # Check if they fit in INTEGER range
                                int_values = [int(val) for val in decimal_values]
                                if all(min_int <= val <= max_int for val in int_values):
                                    col_type = 'INTEGER'
                                else:
                                    col_type = 'BIGINT'
                            else:
                                col_type = 'REAL'
                        except (ValueError, OverflowError):
                            col_type = 'REAL'
                    # Default: VARCHAR with proper sizing
                    else:
                        col_type = f'VARCHAR({varchar_length})'

                    column_info.append({
                        'original_name': col,
                        'column_name': col_name,
                        'postgresql_type': col_type,
                        'sample_values': [str(val) for val in sample_values.head(3).tolist()]
                    })
                
                # Convert datetime columns using the detected format for consistency
                for col, fmt in datetime_columns.items():
                    if fmt:
                        # Use the detected format to parse ALL values consistently
                        df_clean[col] = pd.to_datetime(df_clean[col], format=fmt, errors='coerce')
                    else:
                        # Fallback if format detection failed
                        df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce', infer_datetime_format=True)

                # Get sample data for preview (first 10 rows)
                sample_data = df_clean.head(10).to_dict('records')
                
                return Response({
                    "message": f"File uploaded successfully! Review the table structure below. Table will be created in customer database: {customer.cust_db}",
                    "table_info": {
                        "table_name": table_name,
                        "rows_count": len(df_clean),
                        "columns": column_info
                    },
                    "sample_data": sample_data,
                    "file_name": uploaded_file.name,
                    "file_extension": file_extension,
                    "customer_info": {
                        "cust_id": customer.cust_id,
                        "cust_db": customer.cust_db,
                        "customer_name": customer.name
                    }
                }, status=status.HTTP_200_OK)

            except Exception as e:
                return Response(
                    {"error": f"Error processing file: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




class WriteTableToDatabaseView(APIView):
    """
    API view for writing table structure and data to database without using session storage.
    This view receives the file, table structure, and creates the table with data in the user's customer database.
    """
    
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value
        
        str_value = str(value).strip()
        
        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)
                
                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return str(float_val)
                else:
                    # Default behavior: convert to int for phone numbers and other cases
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value
        
        return str_value

    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()
        
        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)
        
        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')
        
        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)
        
        return df_cleaned
    
    def _detect_date_format(self, series):
        """
        Detect the consistent date format for a column by analyzing unambiguous dates.
        Returns the format string that should be used, or None if no format detected.
        """
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return None
        
        # Sample for performance
        sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values
        
        # Try each format and score based on successful conversions
        formats_to_try = [
            '%Y-%m-%d',                  # 2024-01-15
            '%m/%d/%Y',                  # 01/15/2024 (US format - month first)
            '%d/%m/%Y',                  # 15/01/2024 (European format - day first)
            '%m-%d-%Y',                  # 01-15-2024 (US format - month first)
            '%d-%m-%Y',                  # 15-01-2024 (European format - day first)
            '%Y/%m/%d',                  # 2024/01/15
            '%Y%m%d',                    # 20240115
            '%d.%m.%Y',                  # 15.01.2024
            '%Y-%m-%d %H:%M:%S',         # 2024-01-15 14:30:00
            '%m/%d/%Y %H:%M:%S',         # 01/15/2024 14:30:00
            '%d/%m/%Y %H:%M:%S',         # 15/01/2024 14:30:00
            '%m-%d-%Y %H:%M:%S',         # 01-15-2024 14:30:00
            '%d-%m-%Y %H:%M:%S',         # 15-01-2024 14:30:00
            '%Y-%m-%d %H:%M:%S.%f',      # 2024-01-15 14:30:00.123456
            '%m-%d-%Y %H:%M:%S.%f',      # 01-15-2024 14:30:00.123456
            '%d-%m-%Y %H:%M:%S.%f',      # 15-01-2024 14:30:00.123456
            '%m/%d/%Y %H:%M:%S.%f',      # 01/15/2024 14:30:00.123456
            '%d/%m/%Y %H:%M:%S.%f',      # 15/01/2024 14:30:00.123456
        ]
        
        best_format = None
        best_score = 0
        
        for fmt in formats_to_try:
            try:
                converted = pd.to_datetime(sample, format=fmt, errors='coerce')
                success_count = converted.notna().sum()
                
                # Calculate success rate
                if success_count > best_score:
                    best_score = success_count
                    best_format = fmt
            except (ValueError, TypeError):
                continue
        
        # Return format if at least 60% success rate
        if best_score / len(sample) > 0.6:
            return best_format
        
        return None
    
    def _convert_datetime_value(self, value, column_type):
        """
        Convert datetime value to appropriate format for database insertion.
        Handles both DATE and TIMESTAMP types.
        """
        if pd.isna(value) or value == '' or value is None:
            return None
        
        try:
            # Convert to datetime using pandas
            dt = pd.to_datetime(value, errors='coerce', infer_datetime_format=True)
            
            if pd.isna(dt):
                return None
            
            # Convert pandas Timestamp to Python datetime
            if isinstance(dt, pd.Timestamp):
                dt = dt.to_pydatetime()
            
            # Return the datetime object (PostgreSQL will handle the conversion)
            return dt
        except Exception:
            return None
    
    def post(self, request):
        try:
            # Get the uploaded file
            if 'file' not in request.FILES:
                return Response(
                    {"error": "No file provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            uploaded_file = request.FILES['file']
            table_name = request.data.get('table_name')
            columns_data = request.data.get('columns', [])
            scope = request.data.get('scope')
            schema = request.data.get('schema')

      
            
            # Parse columns if it's a JSON string
            if isinstance(columns_data, str):
                import json
                try:
                    columns = json.loads(columns_data)
                except json.JSONDecodeError:
                    return Response(
                        {"error": "Invalid columns data format."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                columns = columns_data
            
            if not table_name or not columns:
                return Response(
                    {"error": "Table name and columns are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Determine target schema based on scope
            if scope == 'global':
                target_schema = 'GENERAL'
            else:  # local scope
                target_schema = schema

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id  # user.cust_id is already the Customer object
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            file_extension = uploaded_file.name.lower().split('.')[-1]
            
            # Read file into pandas DataFrame
            # Don't use dtype=str to preserve datetime types
            uploaded_file.seek(0)
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
            elif file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Remove empty rows from DataFrame
            df = self._remove_empty_rows(df)
            
            # Clean DataFrame
            df_clean = df.where(pd.notnull(df), None)
            
            # Convert datetime columns using detected format for consistency
            # Check each column to see if it's a date/timestamp type
            for col in columns:
                original_name = col.get('original_name')
                col_type = col.get('postgresql_type', '').upper()
                
                # If this column is DATE or TIMESTAMP type, convert it consistently
                if col_type in ['DATE', 'TIMESTAMP', 'TIMESTAMP WITHOUT TIME ZONE', 'TIMESTAMP WITH TIME ZONE']:
                    if original_name in df_clean.columns:
                        # Detect format and convert
                        detected_format = self._detect_date_format(df_clean[original_name])
                        if detected_format:
                            df_clean[original_name] = pd.to_datetime(df_clean[original_name], format=detected_format, errors='coerce')
                        else:
                            df_clean[original_name] = pd.to_datetime(df_clean[original_name], errors='coerce', infer_datetime_format=True)
            
            # Build SQL for creating table
            column_definitions = []
            
            # Add system fields first
            column_definitions.append('__id SERIAL PRIMARY KEY')
            
            for col in columns:
                col_name = col.get('column_name')
                col_type = col.get('postgresql_type')
                if col_name and col_type:
                    column_definitions.append(f'"{col_name}" {col_type}')
            
            # Add is_active field at the end
            column_definitions.append('__active BOOLEAN DEFAULT TRUE')
            
            if not column_definitions:
                return Response(
                    {"error": "No valid column definitions provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            create_table_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(column_definitions)});'
            
            # Use transaction.on_commit to avoid transaction conflicts
            from django.db import transaction
            
            # Prepare data for table creation (outside transaction)
            # Filter out system fields (__id and is_active) from insert columns since they're auto-generated
            insert_columns = [col for col in columns if col.get('column_name') not in ['__id', 'is_active']]
            column_names = [col.get('column_name') for col in insert_columns]
            original_column_names = [col.get('original_name') for col in insert_columns]
            placeholders = ', '.join(['%s'] * len(column_names))
            insert_sql = f'INSERT INTO "{table_name}" ({", ".join(column_names)}) VALUES ({placeholders})'
            
            data_tuples = []
            for _, row in df_clean.iterrows():
                row_data = []
                for i, col_name in enumerate(column_names):
                    original_col_name = original_column_names[i]
                    col_type = insert_columns[i].get('postgresql_type', '') if i < len(insert_columns) else None
                    
                    # Handle is_active field specially
                    if col_name == 'is_active':
                        row_data.append(True)  # Default to True for is_active
                    else:
                        # Try to get the value using the original column name
                        if original_col_name in df_clean.columns:
                            val = row[original_col_name]
                        else:
                            # If original column name not found, try the database column name
                            val = row[col_name] if col_name in df_clean.columns else None
                        
                        if pd.isna(val):
                            row_data.append(None)
                        # Handle date/timestamp columns
                        elif col_type and col_type.upper() in ['DATE', 'TIMESTAMP', 'TIMESTAMP WITHOUT TIME ZONE', 'TIMESTAMP WITH TIME ZONE']:
                            row_data.append(self._convert_datetime_value(val, col_type))
                        elif isinstance(val, pd.Timestamp):
                            row_data.append(val.to_pydatetime())
                        else:
                            # Apply scientific notation fix for phone numbers and numeric columns before storing
                            if (original_col_name and any(keyword in original_col_name.lower() for keyword in ['phone', 'mobile', 'tel', 'contact'])) or \
                               (col_type and col_type.upper() in ['BIGINT', 'INTEGER', 'SMALLINT', 'NUMERIC', 'DECIMAL']):
                                val = self._fix_scientific_notation(val, col_type)
                            row_data.append(val)
                data_tuples.append(tuple(row_data))
            
            def create_table_in_customer_db():
                # Connect to customer's database instead of default database
                customer_db_config = {
                    'ENGINE': 'django.db.backends.postgresql',
                    'NAME': customer.cust_db,
                    'USER': settings.DATABASES['default']['USER'],
                    'PASSWORD': settings.DATABASES['default']['PASSWORD'],
                    'HOST': settings.DATABASES['default']['HOST'],
                    'PORT': settings.DATABASES['default']['PORT'],
                }
                
                # Create connection to customer database
                customer_connection = psycopg2.connect(
                    host=customer_db_config['HOST'],
                    port=customer_db_config['PORT'],
                    database=customer_db_config['NAME'],
                    user=customer_db_config['USER'],
                    password=customer_db_config['PASSWORD']
                )
                
                # Set autocommit after connection is established
                customer_connection.autocommit = True
                
                with customer_connection.cursor() as cursor:
                    # Set search path to target schema
                    cursor.execute(f'SET search_path TO "{target_schema}";')
                    
                    # Drop table if exists (in target schema)
                    cursor.execute(f'DROP TABLE IF EXISTS "{target_schema}"."{table_name}";')
                    
                    # Create table in target schema
                    create_table_sql_with_schema = f'CREATE TABLE IF NOT EXISTS "{target_schema}"."{table_name}" ({", ".join(column_definitions)});'
                    cursor.execute(create_table_sql_with_schema)
                    
                    # Insert data (in target schema)
                    insert_sql_with_schema = f'INSERT INTO "{target_schema}"."{table_name}" ({", ".join(column_names)}) VALUES ({placeholders})'
                    cursor.executemany(insert_sql_with_schema, data_tuples)

                    # Create the activity_log table if it does not exist
                    create_log_table_sql = '''
                        CREATE TABLE IF NOT EXISTS "GENERAL"."activity_log" (
                            table_name VARCHAR(100) NOT NULL,
                            created_by VARCHAR(100),
                            created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            records_count INTEGER NOT NULL,
                            modified_by VARCHAR(100),
                            modified_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            description TEXT
                        );
                    '''
                    cursor.execute(create_log_table_sql)

                    # Insert a log entry for the newly created table
                    insert_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_log_sql,
                        (f'{target_schema}.{table_name}', user.email, len(data_tuples), None, f"Table '{table_name}' created in {target_schema} schema with {len(data_tuples)} records")
                    )
                
                customer_connection.close()
            
            # Execute table creation after transaction commits
            transaction.on_commit(create_table_in_customer_db)
            
            return Response({
                "message": f"Table '{table_name}' will be created in customer database '{customer.cust_db}' in the {target_schema} schema with {len(columns)} columns and {len(data_tuples)} rows.",
                "table_name": table_name,

                "columns": [col.get('column_name') for col in columns],
                "rows_inserted": len(data_tuples),
                "customer_info": {
                    "cust_id": customer.cust_id,
                    "cust_db": customer.cust_db,
                    "customer_name": customer.name
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error creating table: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ListUploadedTablesView(APIView):
    """
    API view to list tables ending with '__' from the user's customer database GENERAL schema.
    Requires JWT authentication.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, project_id):
        try:
            # Get the authenticated user from the request
            user = request.user
            
            # Check if user is associated with a customer
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            customer = user.cust_id  # user.cust_id is already the Customer object
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # # Set search path to GENERAL schema
                # cursor.execute('SET search_path TO "GENERAL";')
                
                # Query for tables ending with '__' in GENERAL schema (global)
                cursor.execute("""
                    SELECT table_name
                    FROM information_schema.tables
                    WHERE table_schema = 'GENERAL'
                    AND table_name LIKE '%__'             -- Ends with two underscores
                    AND RIGHT(table_name, 2) = '__'       -- Enforces that the two underscores are at the end
                    ORDER BY table_name;
                """)
                tables_query = cursor.fetchall()

                global_tables = []
                for row in tables_query:
                    global_tables.append({
                        'table_name': row[0],
                        'schema': 'GENERAL'
                    })

                # Query for schemas containing the project_id in their name (local)
                cursor.execute("""
                    SELECT schema_name
                    FROM information_schema.schemata
                    WHERE schema_name ILIKE %s
                    ORDER BY schema_name;
                """, ['%' + project_id + '%'])
                schemas_query = cursor.fetchall()
                schemas = [schema[0] for schema in schemas_query]

                # If there is exactly one schema, fetch table names ending with '__' from that schema
                local_tables = []
                if len(schemas) == 1:
                    schema_name = schemas[0] # schema name is FAD_M01
                    cursor.execute("""
                        SELECT table_name
                        FROM information_schema.tables
                        WHERE table_schema = %s
                        AND RIGHT(table_name, 2) = '__'
                        ORDER BY table_name;
                    """, [schema_name])
                    local_tables_query = cursor.fetchall()
                    
                    for row in local_tables_query:
                        local_tables.append({
                            'table_name': row[0],
                            'schema': schema_name
                        })

            customer_connection.close()

            
            return Response({
                "message": f"Found {len(global_tables)} tables ending with '__' in GENERAL schema.",
                "global_tables": global_tables,
                "local_tables": local_tables,
                "schema": schema_name,
                "customer_info": {
                    "cust_id": customer.cust_id,
                    "cust_db": customer.cust_db,
                    "customer_name": customer.name
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error listing tables: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class GetTableDataView(APIView):
    """
    API view to retrieve records from a specific table in the user's customer database with pagination.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def _build_where_clause(self, filters, columns):
        """
        Build WHERE clause and parameters from filters list.
        Filters format: [{"column": "name", "operator": "=", "value": "John"}, ...]
        Supported operators: =, !=, >, <, >=, <=, LIKE, ILIKE, IN, NOT IN
        """
        if not filters:
            return "", []
        
        where_conditions = []
        where_params = []
        column_names = [col[0] for col in columns]
        
        for filter_item in filters:
            column = filter_item.get('column', '')
            operator = filter_item.get('operator', '=')
            value = filter_item.get('value', '')
            
            # Validate column exists
            if column not in column_names:
                continue
                
            # Handle different operators
            if operator == '=':
                where_conditions.append(f'"{column}" = %s')
                where_params.append(value)
            elif operator == '!=':
                where_conditions.append(f'"{column}" != %s')
                where_params.append(value)
            elif operator == '>':
                where_conditions.append(f'"{column}" > %s')
                where_params.append(value)
            elif operator == '<':
                where_conditions.append(f'"{column}" < %s')
                where_params.append(value)
            elif operator == '>=':
                where_conditions.append(f'"{column}" >= %s')
                where_params.append(value)
            elif operator == '<=':
                where_conditions.append(f'"{column}" <= %s')
                where_params.append(value)
            elif operator == 'LIKE':
                where_conditions.append(f'"{column}" LIKE %s')
                where_params.append(f'%{value}%')
            elif operator == 'ILIKE':
                where_conditions.append(f'"{column}" ILIKE %s')
                where_params.append(f'%{value}%')
            elif operator == 'IN':
                # Handle both list and comma-separated string
                if isinstance(value, list) and value:
                    placeholders = ','.join(['%s'] * len(value))
                    where_conditions.append(f'"{column}" IN ({placeholders})')
                    where_params.extend(value)
                elif isinstance(value, str) and value:
                    # Split comma-separated string into list
                    values_list = [v.strip() for v in value.split(',') if v.strip()]
                    if values_list:
                        placeholders = ','.join(['%s'] * len(values_list))
                        where_conditions.append(f'"{column}" IN ({placeholders})')
                        where_params.extend(values_list)
            elif operator == 'NOT IN':
                # Handle both list and comma-separated string
                if isinstance(value, list) and value:
                    placeholders = ','.join(['%s'] * len(value))
                    where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                    where_params.extend(value)
                elif isinstance(value, str) and value:
                    # Split comma-separated string into list
                    values_list = [v.strip() for v in value.split(',') if v.strip()]
                    if values_list:
                        placeholders = ','.join(['%s'] * len(values_list))
                        where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                        where_params.extend(values_list)
            elif operator == 'IS NULL':
                where_conditions.append(f'"{column}" IS NULL')
            elif operator == 'IS NOT NULL':
                where_conditions.append(f'"{column}" IS NOT NULL')
            elif operator == 'MISMATCH':
                # Handle MISMATCH operator - value contains JSON string with mismatch pairs
                try:
                    import json
                    mismatch_data = json.loads(value) if isinstance(value, str) else value
                    if isinstance(mismatch_data, list) and len(mismatch_data) > 0:
                        # Get the first item to determine related column name
                        first_item = mismatch_data[0]
                        related_column = None
                        for key in first_item.keys():
                            if key != column:
                                related_column = key
                                break
                        
                        if related_column and related_column in column_names:
                            # Build OR conditions for each mismatch pair
                            mismatch_conditions = []
                            for mismatch_pair in mismatch_data:
                                mismatch_conditions.append(
                                    f'("{column}" = %s AND "{related_column}" = %s)'
                                )
                                where_params.append(mismatch_pair.get(column))
                                where_params.append(mismatch_pair.get(related_column))
                            
                            # Join with OR
                            where_conditions.append(f'({" OR ".join(mismatch_conditions)})')
                except (json.JSONDecodeError, AttributeError, KeyError) as e:
                    # If JSON parsing fails, skip this filter
                    continue
        
        if where_conditions:
            return f' WHERE {" AND ".join(where_conditions)}', where_params
        return "", []
    
    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            project_id = request.data.get('project_id')
            schema = request.data.get('schema')
            page = int(request.data.get('page', 1))
            page_size = int(request.data.get('page_size', 100))
            sort_column = request.data.get('sort_column', '')
            sort_direction = request.data.get('sort_direction', 'asc')  # 'asc' or 'desc'
            filters = request.data.get('filters', [])  # List of filter objects

            
            if not table_name:
                return Response(
                    {"error": "Table name is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate pagination parameters
            if page < 1:
                page = 1
            if page_size < 1 or page_size > 1000:  # Limit page size to prevent abuse
                page_size = 100
            
            # Validate sort parameters
            if sort_direction not in ['asc', 'desc']:
                sort_direction = 'asc'

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id  # user.cust_id is already the Customer object
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                
                # First, get column information including is_nullable
                cursor.execute("""
                    SELECT column_name, data_type, is_nullable
                    FROM information_schema.columns 
                    WHERE table_schema = %s 
                    AND table_name = %s
                    ORDER BY ordinal_position;
                """, (schema, table_name))
                
                columns = cursor.fetchall()
                if not columns:
                    return Response(
                        {"error": f"Table '{table_name}' not found or has no columns."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Build WHERE clause for filtering
                where_clause, where_params = self._build_where_clause(filters, columns)
                
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'
                
                # Get total count of records for pagination info (with filters)
                count_query = f'SELECT COUNT(*) FROM {table_reference}{where_clause};'
                cursor.execute(count_query, where_params)
                total_records = cursor.fetchone()[0]
                
                # Calculate pagination
                offset = (page - 1) * page_size
                total_pages = (total_records + page_size - 1) // page_size  # Ceiling division
                
                # Build ORDER BY clause for sorting
                order_by_clause = ""
                if sort_column and sort_column in [col[0] for col in columns]:
                    # Validate that the sort column exists in the table
                    order_by_clause = f' ORDER BY "{sort_column}" {sort_direction.upper()}'
                
                

                # Get paginated data from the table with filtering and sorting
                query = f'SELECT * FROM {table_reference}{where_clause}{order_by_clause} LIMIT %s OFFSET %s;'
                cursor.execute(query, where_params + [page_size, offset])
                rows = cursor.fetchall()
                
                # Convert rows to list of dictionaries
                data = []
                column_names = [col[0] for col in columns]
                
                for row in rows:
                    row_dict = {}
                    for i, value in enumerate(row):
                        row_dict[column_names[i]] = value
                    data.append(row_dict)
                
                # Format date/timestamp columns based on user's preferred date format
                user_strftime_format = convert_user_date_format_to_strftime(user.date_format)
                data = format_date_columns(data, columns, user_strftime_format)
                
                # Check if table has primary key
                cursor.execute("""
                    SELECT column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu 
                        ON tc.constraint_name = kcu.constraint_name
                    WHERE tc.table_schema = %s 
                        AND tc.table_name = %s 
                        AND tc.constraint_type = 'PRIMARY KEY'
                    ORDER BY kcu.ordinal_position;
                """, (schema, table_name))
                
                pk_columns = cursor.fetchall()
                has_primary_key = len(pk_columns) > 0
            
            customer_connection.close()
            
            # Get user's preferred file format
            user_file_format = user.file_format

            
            return Response({
                "message": f"Retrieved {len(data)} records from table '{table_name}' (page {page} of {total_pages}).",
                "table_name": table_name,
                "columns": [{"name": col[0], "type": col[1], "is_nullable": col[2]} for col in columns],
                "data": data,
                "pagination": {
                    "current_page": page,
                    "page_size": page_size,
                    "total_pages": total_pages,
                    "total_records": total_records,
                    "has_next": page < total_pages,
                    "has_previous": page > 1,
                    "showing_start": offset + 1 if total_records > 0 else 0,
                    "showing_end": min(offset + page_size, total_records)
                },
                "sorting": {
                    "sort_column": sort_column,
                    "sort_direction": sort_direction
                },
                "filters": filters,
                "has_primary_key": has_primary_key,
                "primary_key_columns": [col[0] for col in pk_columns] if has_primary_key else [],
                "customer_info": {
                    "cust_id": customer.cust_id,
                    "cust_db": customer.cust_db,
                    "customer_name": customer.name
                },
                "user_file_format": user_file_format
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error retrieving table data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class GetDistinctValuesView(APIView):
    """
    API view to get distinct values for a specific column in a table.
    Used for populating filter dropdowns.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            column_name = request.data.get('column_name')
            schema = request.data.get('schema')
            limit = int(request.data.get('limit', 1000))  # Limit distinct values returned
            
            if not table_name or not column_name:
                return Response(
                    {"error": "Table name and column name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Check if column exists
                cursor.execute("""
                    SELECT column_name, data_type
                    FROM information_schema.columns 
                    WHERE table_schema = %s 
                    AND table_name = %s
                    AND column_name = %s
                """, (schema, table_name, column_name))
                
                column_info = cursor.fetchone()
                if not column_info:
                    return Response(
                        {"error": f"Column '{column_name}' not found in table '{table_name}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"' if schema else f'"{table_name}"'
                
                # Get distinct values
                cursor.execute(f'SELECT DISTINCT "{column_name}" FROM {table_reference} WHERE "{column_name}" IS NOT NULL ORDER BY "{column_name}" LIMIT %s;', (limit,))
                distinct_values = [row[0] for row in cursor.fetchall()]
                
                # Convert to string for JSON serialization
                distinct_values = [str(value) if value is not None else '' for value in distinct_values]
            
            customer_connection.close()
            
            return Response({
                "column_name": column_name,
                "data_type": column_info[1],
                "distinct_values": distinct_values,
                "total_count": len(distinct_values)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error retrieving distinct values: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PreviewTableDataView(APIView):
    """
    API view to preview data from uploaded file for table insertion.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value
        
        str_value = str(value).strip()
        
        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)
                
                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return f"{float_val:.10f}".rstrip('0').rstrip('.')
                else:
                    # Default: convert to int (for backward compatibility with phone numbers)
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value
        
        return str_value


    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()
        
        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)
        
        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')
        
        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)

        df_cleaned = df_cleaned.drop_duplicates()
        
        return df_cleaned
    
    def post(self, request):
        try:
            # Get the uploaded file
            if 'file' not in request.FILES:
                return Response(
                    {"error": "No file provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            uploaded_file = request.FILES['file']
            table_name = request.data.get('table_name')
            schema = request.data.get('schema') # Added schema extraction
            
            if not table_name:
                return Response(
                    {"error": "Table name is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information from authenticated user
            user = request.user 
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            customer = user.cust_id
            
            file_extension = uploaded_file.name.lower().split('.')[-1]
            
            # Read file into pandas DataFrame
            # Don't use dtype=str to preserve datetime types
            uploaded_file.seek(0)
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
            elif file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Remove empty rows from DataFrame
            df = self._remove_empty_rows(df)
            
            # Clean DataFrame
            df_clean = df.where(pd.notnull(df), None)
            
            # Fix scientific notation for all columns that might contain it
            for col in df_clean.columns:
                # Check if any value in this column contains scientific notation
                has_scientific_notation = False
                for val in df_clean[col].dropna():
                    if isinstance(val, str) and 'E' in val.upper():
                        has_scientific_notation = True
                        break
                
                if has_scientific_notation:
                    # Convert scientific notation back to regular numbers
                    df_clean[col] = df_clean[col].apply(lambda x: self._fix_scientific_notation(x) if pd.notna(x) else x)
            
            # Convert to list of dictionaries for JSON serialization
            sample_data = df_clean.head(10).to_dict('records')
            
            return Response({
                "message": f"File preview successful. Found {len(df_clean)} rows of data.",
                "sample_data": sample_data,
                "total_rows": len(df_clean),
                "columns": list(df_clean.columns),
                "table_name": table_name
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error previewing data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class UploadTableDataView(APIView):
    """
    API view to upload data to an existing table in the customer's database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value
        
        str_value = str(value).strip()
        
        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)
                
                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return f"{float_val:.10f}".rstrip('0').rstrip('.')
                else:
                    # Default: convert to int (for backward compatibility with phone numbers)
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value
        
        return str_value

    def _remove_duplicates_from_new_data(self, new_df):
        """
        Remove duplicates from new data
        """
        return new_df.drop_duplicates()
    
    def _detect_date_format(self, series):
        """
        Detect the consistent date format for a column by analyzing unambiguous dates.
        Returns the format string that should be used, or None if no format detected.
        """
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return None
        
        # Sample for performance
        sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values
        
        # Try each format and score based on successful conversions
        formats_to_try = [
            '%Y-%m-%d',                  # 2024-01-15
            '%m/%d/%Y',                  # 01/15/2024 (US format - month first)
            '%d/%m/%Y',                  # 15/01/2024 (European format - day first)
            '%m-%d-%Y',                  # 01-15-2024 (US format - month first)
            '%d-%m-%Y',                  # 15-01-2024 (European format - day first)
            '%Y/%m/%d',                  # 2024/01/15
            '%Y%m%d',                    # 20240115
            '%d.%m.%Y',                  # 15.01.2024
            '%Y-%m-%d %H:%M:%S',         # 2024-01-15 14:30:00
            '%m/%d/%Y %H:%M:%S',         # 01/15/2024 14:30:00
            '%d/%m/%Y %H:%M:%S',         # 15/01/2024 14:30:00
            '%m-%d-%Y %H:%M:%S',         # 01-15-2024 14:30:00
            '%d-%m-%Y %H:%M:%S',         # 15-01-2024 14:30:00
            '%Y-%m-%d %H:%M:%S.%f',      # 2024-01-15 14:30:00.123456
            '%m-%d-%Y %H:%M:%S.%f',      # 01-15-2024 14:30:00.123456
            '%d-%m-%Y %H:%M:%S.%f',      # 15-01-2024 14:30:00.123456
            '%m/%d/%Y %H:%M:%S.%f',      # 01/15/2024 14:30:00.123456
            '%d/%m/%Y %H:%M:%S.%f',      # 15/01/2024 14:30:00.123456
        ]
        
        best_format = None
        best_score = 0
        
        for fmt in formats_to_try:
            try:
                converted = pd.to_datetime(sample, format=fmt, errors='coerce')
                success_count = converted.notna().sum()
                
                # Calculate success rate
                if success_count > best_score:
                    best_score = success_count
                    best_format = fmt
            except (ValueError, TypeError):
                continue
        
        # Return format if at least 60% success rate
        if best_score / len(sample) > 0.6:
            return best_format
        
        return None
    
    def _convert_datetime_value(self, value, column_type):
        """
        Convert datetime value to appropriate format for database insertion.
        Handles both DATE and TIMESTAMP types.
        """
        if pd.isna(value) or value == '' or value is None:
            return None
        
        try:
            # Convert to datetime using pandas
            dt = pd.to_datetime(value, errors='coerce', infer_datetime_format=True)
            
            if pd.isna(dt):
                return None
            
            # Convert pandas Timestamp to Python datetime
            if isinstance(dt, pd.Timestamp):
                dt = dt.to_pydatetime()
            
            # Return the datetime object (PostgreSQL will handle the conversion)
            return dt
        except Exception:
            return None
        

    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows and duplicate rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()
        
        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)
        
        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')
        
        # Remove duplicate rows (keep first occurrence)
        df_cleaned = df_cleaned.drop_duplicates()
        
        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)
        return df_cleaned
    
    def get_column_info(self, cursor, table_name, schema):
        """
        Get comprehensive column information from the database schema.
        Returns a dictionary with column names as keys and their details as values.
        """
        cursor.execute("""
            SELECT column_name, data_type, character_maximum_length, numeric_precision, 
                   numeric_scale, datetime_precision, is_nullable
            FROM information_schema.columns 
            WHERE table_schema = %s 
            AND table_name = %s 
            AND column_name NOT IN ('__active', '__id')
            ORDER BY ordinal_position;
        """, (schema, table_name))
        
        column_info = {}
        for col_name, data_type, char_max_length, numeric_precision, numeric_scale, datetime_precision, is_nullable in cursor.fetchall():
            column_info[col_name] = {
                'data_type': data_type,
                'char_max_length': char_max_length,
                'numeric_precision': numeric_precision,
                'numeric_scale': numeric_scale,
                'datetime_precision': datetime_precision,
                'is_nullable': is_nullable == 'YES'
            }
        
        return column_info
    
    def post(self, request):
        try:
            # Get the uploaded file
            if 'file' not in request.FILES:
                return Response(
                    {"error": "No file provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            uploaded_file = request.FILES['file']
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            
            if not table_name:
                return Response(
                    {"error": "Table name is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            file_extension = uploaded_file.name.lower().split('.')[-1]
            
            # Read file into pandas DataFrame
            # Don't use dtype=str to preserve datetime types
            uploaded_file.seek(0)
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
            elif file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Remove empty rows from DataFrame
            df = self._remove_empty_rows(df)
            
            # Clean DataFrame
            df_clean = df.where(pd.notnull(df), None)
            
            # Check for reserved field names in the uploaded file
            reserved_fields = ['__id', '__active']
            found_reserved_fields = []
            for col in df_clean.columns:
                col_lower = col.lower().strip()
                if col_lower in reserved_fields or col_lower == 'id':
                    found_reserved_fields.append(col)
            
            if found_reserved_fields:
                return Response(
                    {"error": f"The following field names are reserved and cannot be used: {', '.join(found_reserved_fields)}. Please rename these columns in your file and try again."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Note: Scientific notation fix will be applied later based on actual database column types
            
            # Get table structure to match columns and detect datetime columns
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            # Don't set autocommit = True to allow savepoints
            customer_connection.autocommit = False
            
            with customer_connection.cursor() as cursor:
                # Get table columns (excluding __active and __id as they will be added automatically)
                cursor.execute("""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_schema = %s 
                    AND table_name = %s 
                    AND column_name NOT IN ('__active', '__id')
                    ORDER BY ordinal_position;
                """, (schema, table_name))
                
                table_columns = cursor.fetchall()
                
                if not table_columns:
                    customer_connection.close()
                    return Response(
                        {"error": f"Table '{table_name}' not found in schema '{schema}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Get comprehensive column information
                column_info = self.get_column_info(cursor, table_name, schema)
                
                # Validate column names match between file and database table
                db_column_names = [col[0].lower() for col in table_columns]
                file_column_names = [col.lower().strip() for col in df_clean.columns]
                
                # Find missing columns in file (database columns not found in file)
                missing_in_file = []
                for db_col in db_column_names:
                    if db_col not in file_column_names:
                        missing_in_file.append(db_col)
                
                # Find extra columns in file (file columns not found in database)
                extra_in_file = []
                for file_col in file_column_names:
                    if file_col not in db_column_names:
                        extra_in_file.append(file_col)
                
                # If there are missing or extra columns, return detailed error
                if missing_in_file or extra_in_file:
                    error_message = "Column mismatch between uploaded file and database table:\n"
                    
                    if missing_in_file:
                        error_message += f"Missing columns in file: {', '.join(missing_in_file)}\n"
                    
                    if extra_in_file:
                        error_message += f"Extra columns in file (not in database): {', '.join(extra_in_file)}\n"
                    
                    error_message += f"\nDatabase table columns: {', '.join(db_column_names)}\n"
                    error_message += f"File columns: {', '.join(file_column_names)}"
                    
                    customer_connection.close()
                    return Response(
                        {"error": error_message},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Map file columns to table columns (case-insensitive)
                column_mapping = {}
                file_columns = [col.lower() for col in df_clean.columns]
                
                for db_col, db_type in table_columns:
                    # Try to find matching column in file
                    matching_col = None
                    for file_col in df_clean.columns:
                        if file_col.lower() == db_col.lower():
                            matching_col = file_col
                            break
                    
                    if matching_col:
                        column_mapping[db_col] = matching_col
                
                if not column_mapping:
                    customer_connection.close()
                    return Response(
                        {"error": "No matching columns found between file and table."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Convert datetime columns using detected format for consistency
                for db_col, file_col in column_mapping.items():
                    db_col_info = column_info.get(db_col, {})
                    db_data_type = db_col_info.get('data_type', '').lower()
                    
                    # If this column is date/timestamp type, convert it consistently
                    if db_data_type in ['date', 'timestamp', 'timestamp without time zone', 'timestamp with time zone']:
                        # Detect format and convert
                        detected_format = self._detect_date_format(df_clean[file_col])
                        if detected_format:
                            df_clean[file_col] = pd.to_datetime(df_clean[file_col], format=detected_format, errors='coerce')
                        else:
                            df_clean[file_col] = pd.to_datetime(df_clean[file_col], errors='coerce', infer_datetime_format=True)
                
                # Use the original data (no truncation needed)
                processed_df = df_clean
                
                # Prepare data for insertion
                insert_columns = list(column_mapping.keys())
                placeholders = ', '.join(['%s'] * (len(insert_columns) + 1))  # +1 for __active
                table_reference = f'"{schema}"."{table_name}"'
                insert_sql = f'INSERT INTO {table_reference} ({", ".join(insert_columns)}, __active) VALUES ({placeholders})'
                
                data_tuples = []
                for _, row in processed_df.iterrows():
                    row_data = []
                    for db_col in insert_columns:
                        file_col = column_mapping[db_col]
                        val = row[file_col]
                        
                        if pd.isna(val):
                            row_data.append(None)
                        else:
                            # Get database column information
                            db_col_info = column_info.get(db_col, {})
                            db_data_type = db_col_info.get('data_type', '').lower()
                            
                            # Handle date/timestamp columns
                            if db_data_type in ['date', 'timestamp', 'timestamp without time zone', 'timestamp with time zone']:
                                row_data.append(self._convert_datetime_value(val, db_data_type))
                            elif isinstance(val, pd.Timestamp):
                                row_data.append(val.to_pydatetime())
                            else:
                                # Apply scientific notation fix based on database column type
                                # Check if this is a numeric column that might have scientific notation
                                if db_data_type and any(numeric_type in db_data_type for numeric_type in ['bigint', 'integer', 'smallint', 'numeric', 'decimal', 'real', 'double']):
                                    val = self._fix_scientific_notation(val, db_data_type)
                                
                                # Send raw string value - let PostgreSQL handle type validation
                                row_data.append(str(val))
                    
                    # Add __active = True for all rows
                    row_data.append(True)
                    data_tuples.append(tuple(row_data))

           
                # Start transaction for all insertions
                cursor.execute("BEGIN TRANSACTION")
                
                # Use batch processing for better performance
                successful_inserts = 0
                failed_inserts = 0
                failed_records = []
                batch_size = 100  # Process in batches of 100 records
                
                # Adjust batch size based on total records for optimal performance
                total_records = len(data_tuples)
                if total_records > 1000:
                    batch_size = 200
                elif total_records > 5000:
                    batch_size = 500
                elif total_records > 10000:
                    batch_size = 1000
                
                # Process all records in batches using savepoints
                for batch_start in range(0, len(data_tuples), batch_size):
                    batch_end = min(batch_start + batch_size, len(data_tuples))
                    batch_data = data_tuples[batch_start:batch_end]
                    
                    # Create a savepoint for this batch
                    savepoint_name = f"batch_{batch_start}"
                    cursor.execute(f"SAVEPOINT {savepoint_name}")
                    
                    try:
                        # Try to insert the entire batch
                        cursor.executemany(insert_sql, batch_data)
                        successful_inserts += len(batch_data)
                        # Release the savepoint since batch was successful
                        cursor.execute(f"RELEASE SAVEPOINT {savepoint_name}")
                        
                    except Exception as batch_error:
                        # If batch fails, rollback to savepoint and process individually
                        cursor.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
                        
                        # Process each record in the failed batch individually
                        for i, data_tuple in enumerate(batch_data):
                            # Create a savepoint for each individual record
                            record_savepoint = f"record_{batch_start}_{i}"
                            cursor.execute(f"SAVEPOINT {record_savepoint}")
                            
                            try:
                                cursor.execute(insert_sql, data_tuple)
                                successful_inserts += 1
                                # Release the savepoint since record was successful
                                cursor.execute(f"RELEASE SAVEPOINT {record_savepoint}")
                                
                            except Exception as insert_error:
                                # Rollback to the record savepoint
                                cursor.execute(f"ROLLBACK TO SAVEPOINT {record_savepoint}")
                                failed_inserts += 1
                                
                                # Get the row data for error reporting
                                actual_row_index = batch_start + i
                                row_data = processed_df.iloc[actual_row_index]
                                
                                # Get the first column value for identification
                                first_column_name = processed_df.columns[0]
                                first_column_value = row_data[first_column_name] if first_column_name in row_data else 'N/A'
                                
                                failed_record_info = {
                                    'row_number': actual_row_index + 2,  # +2 because first row is header, so actual data starts from row 2
                                    'first_column_value': str(first_column_value),
                                    'error': str(insert_error),
                                    'data': {col: row_data[col] for col in processed_df.columns}
                                }
                                failed_records.append(failed_record_info)
                
                # Check if any records failed - if so, rollback entire transaction
                if failed_inserts > 0:
                    cursor.execute("ROLLBACK")
                    customer_connection.close()
                    
                    # Prepare error response with all failed records
                    error_details = []
                    for failed_record in failed_records:
                        error_details.append(f"Row {failed_record['row_number']} ({failed_record['first_column_value']}): {failed_record['error'].split('LINE 1:')[0].strip()}")
                    
                    error_message = f"Transaction rolled back. {failed_inserts} records failed to insert. "
                    if len(failed_records) > 10:
                        error_message += f"First 10 errors: {'; '.join(error_details[:10])}"
                    else:
                        error_message += f"All errors: {'; '.join(error_details)}"
                    
                    return Response(
                        {"error": error_message, "failed_records": failed_records},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                else:
                    # All records succeeded - commit the transaction
                    cursor.execute("COMMIT")
                    
                    # Log the file upload activity
                    try:
                        insert_activity_log_sql = '''
                            INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                            VALUES (%s, %s, %s, %s, %s)
                        '''
                        cursor.execute(
                            insert_activity_log_sql,
                            (table_name, None, successful_inserts, user.email, f"File uploaded to table '{table_name}' with {successful_inserts} records")
                        )
                        customer_connection.commit()
                    except Exception as log_error:
                        # Log error but don't fail the main operation
                        print(f"Warning: Failed to log activity: {log_error}")
            
            customer_connection.close()
            
            # All records inserted successfully (transaction was committed)
            response_data = {
                "message": f"Successfully uploaded {successful_inserts} rows to table '{table_name}'."
            }
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            error_message = str(e)
            
            # Check if it's a PostgreSQL data type validation error
            if "invalid input syntax" in error_message.lower() or "invalid value" in error_message.lower():
                # Extract column and value information from the error
                if "for type" in error_message.lower():
                    return Response(
                        {"error": f"Data type validation error: {error_message}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
             
            return Response(
                {"error": f"Error uploading data: {error_message}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class CreateTableRecordView(APIView):
    """
    API view to create a new record in a specific table in the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            record_data = request.data.get('record_data')
            
            if not table_name or not record_data:
                return Response(
                    {"error": "Table name and record data are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Build INSERT query
                columns = list(record_data.keys())
                placeholders = ['%s'] * len(columns)
                values = list(record_data.values())
                
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'
                
                quoted_columns = [f'"{col}"' for col in columns]
                insert_query = f"""
                    INSERT INTO {table_reference} ({', '.join(quoted_columns)})
                    VALUES ({', '.join(placeholders)})
                """
                
                cursor.execute(insert_query, values)
                
                # Log the record creation activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, 1, user.email, f"Record saved to table '{table_name}'")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")
            
            customer_connection.close()
            
            return Response({
                "message": f"Record created successfully in table '{table_name}'.",
                "table_name": table_name,
                "created_fields": columns
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            error_message = str(e)
            
            # Check if it's a PostgreSQL data type validation error
            if "invalid input syntax" in error_message.lower() or "invalid value" in error_message.lower():
                # Extract column and value information from the error
                if "for type" in error_message.lower():
                    return Response(
                        {"error": f"Data type validation error: {error_message}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            return Response(
                {"error": f"Error creating record: {error_message}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EditTableRecordView(APIView):
    """
    API view to edit an existing record in a specific table in the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def put(self, request):
        try:
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            record_id = request.data.get('record_id')
            record_data = request.data.get('record_data')
            
            if not table_name or not record_id or not record_data:
                return Response(
                    {"error": "Table name, record ID, and record data are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # First, get the primary key column name
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu 
                    ON tc.constraint_name = kcu.constraint_name
                    WHERE tc.table_name = %s 
                    AND tc.constraint_type = 'PRIMARY KEY'
                    AND tc.table_schema = %s
                """, [table_name, schema])
                
                pk_result = cursor.fetchone()
                if not pk_result:
                    return Response(
                        {"error": f"No primary key found for table '{table_name}'."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                pk_column = pk_result[0]
                
                # Get table column information for validation
                cursor.execute("""
                    SELECT column_name, data_type, character_maximum_length, is_nullable, column_default
                    FROM information_schema.columns 
                    WHERE table_name = %s AND table_schema = %s
                    ORDER BY ordinal_position
                """, [table_name, schema])
                
                column_info = {row[0]: {
                    'data_type': row[1],
                    'max_length': row[2],
                    'is_nullable': row[3] == 'YES',
                    'default': row[4]
                } for row in cursor.fetchall()}
                
                # Validate record data
                validation_errors = []
                columns = list(record_data.keys())
                
    
                
                # Check if primary key is included
                if pk_column not in columns:
                    return Response(
                        {"error": f"Primary key column '{pk_column}' must be included in the update data."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Validate each field
                for col_name, col_value in record_data.items():
                    if col_name == pk_column:
                        continue  # Skip primary key validation
                    
       
                    
                    if col_name not in column_info:
                        validation_errors.append(f"Column '{col_name}' does not exist in table '{table_name}'")
                        continue
                    
                    col_info = column_info[col_name]
                    
                    
                    # Handle NULL/empty values
                    is_null_or_empty = (col_value is None or col_value == '' or str(col_value).strip() == '')
                    
                    
                    if is_null_or_empty:
                        # Check if column allows NULL values
                        if not col_info['is_nullable'] and col_info['default'] is None:
                            validation_errors.append(f"Column '{col_name}' cannot be NULL or empty")
                        # Convert empty strings to None for database
                        record_data[col_name] = None
                        
                        continue
                    
                    # Validate data type for non-null values
                    type_error = self._validate_data_type(col_name, col_value, col_info['data_type'])
                    if type_error:
                        validation_errors.append(type_error)
                        
                    
                    # Validate field length for character types
                    if col_info['max_length'] and col_info['data_type'] in ['character varying', 'character', 'text']:
                        if len(str(col_value)) > col_info['max_length']:
                            validation_errors.append(f"Column '{col_name}' value exceeds maximum length of {col_info['max_length']} characters")
                            
                
                # Return validation errors if any
                if validation_errors:
                    return Response(
                        {"error": "Validation failed", "details": validation_errors},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Build UPDATE query
                set_clauses = []
                values = []
                for col in columns:
                    if col != pk_column:
                        set_clauses.append(f'"{col}" = %s')
                        values.append(record_data[col])
                
                values.append(record_id)  # Add the record ID for WHERE clause
                
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'
                
                update_query = f"""
                    UPDATE {table_reference} 
                    SET {', '.join(set_clauses)}
                    WHERE "{pk_column}" = %s
                """
                
                cursor.execute(update_query, values)
                
                if cursor.rowcount == 0:
                    return Response(
                        {"error": f"No record found with ID '{record_id}' in table '{table_name}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Log the record update activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, 1, user.email, f"Record updated in table '{table_name}' (ID: {record_id})")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")
            
            customer_connection.close()
            
            return Response({
                "message": f"Record updated successfully in table '{table_name}'.",
                "table_name": table_name,
                "record_id": record_id,
                "updated_fields": [col for col in columns if col != pk_column]
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"EditTableRecordView Error: {str(e)}")
            print(f"Traceback: {error_details}")
            return Response(
                {"error": f"Error updating record: {str(e)}", "debug": error_details},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _validate_data_type(self, column_name, value, data_type):
        """
        Validate that a value matches the expected data type for a column.
        Returns None if valid, or an error message if invalid.
        """
        try:
            # Skip validation for None values (handled separately)
            if value is None:
                return None
                
            data_type_lower = data_type.lower()
            
            # Integer types
            if data_type_lower in ['integer', 'bigint', 'smallint', 'serial', 'bigserial']:
                try:
                    # Try to convert to int, handle both string and numeric inputs
                    int(float(str(value)))  # Convert to float first to handle "123.0" -> 123
                except (ValueError, TypeError, OverflowError):
                    return f"Column '{column_name}' must be an integer"
            
            # Numeric/Decimal types
            elif data_type_lower in ['numeric', 'decimal', 'real', 'double precision', 'float']:
                try:
                    float(str(value))
                except (ValueError, TypeError, OverflowError):
                    return f"Column '{column_name}' must be a number"
            # Boolean type
            elif data_type_lower == 'boolean':
                str_value = str(value).lower().strip()
                if str_value not in ['true', 'false', '1', '0', 'yes', 'no', 'on', 'off', 't', 'f']:
                    return f"Column '{column_name}' must be a boolean value (true/false)"
            
            # Date types
            elif data_type_lower == 'date':
                try:
                    from datetime import datetime
                    str_value = str(value).strip()
                    # Try different date formats
                    formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']
                    parsed = False
                    for fmt in formats:
                        try:
                            datetime.strptime(str_value, fmt)
                            parsed = True
                            break
                        except ValueError:
                            continue
                    if not parsed:
                        return f"Column '{column_name}' must be a valid date (YYYY-MM-DD format)"
                except Exception:
                    return f"Column '{column_name}' must be a valid date"
            
            # Timestamp types
            elif data_type_lower in ['timestamp', 'timestamp without time zone', 'timestamp with time zone']:
                try:
                    from datetime import datetime
                    str_value = str(value).strip()
                    # Try different timestamp formats
                    formats = [
                        '%Y-%m-%d %H:%M:%S', 
                        '%Y-%m-%dT%H:%M:%S', 
                        '%Y-%m-%d %H:%M:%S.%f',
                        '%Y-%m-%dT%H:%M:%S.%f',
                        '%m/%d/%Y %H:%M:%S',
                        '%d/%m/%Y %H:%M:%S'
                    ]
                    parsed = False
                    for fmt in formats:
                        try:
                            datetime.strptime(str_value, fmt)
                            parsed = True
                            break
                        except ValueError:
                            continue
                    if not parsed:
                        return f"Column '{column_name}' must be a valid timestamp"
                except Exception:
                    return f"Column '{column_name}' must be a valid timestamp"
            
            # Time types
            elif data_type_lower == 'time':
                try:
                    from datetime import datetime
                    str_value = str(value).strip()
                    formats = ['%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p']
                    parsed = False
                    for fmt in formats:
                        try:
                            datetime.strptime(str_value, fmt)
                            parsed = True
                            break
                        except ValueError:
                            continue
                    if not parsed:
                        return f"Column '{column_name}' must be a valid time (HH:MM:SS format)"
                except Exception:
                    return f"Column '{column_name}' must be a valid time"
            
            # UUID type
            elif data_type_lower == 'uuid':
                try:
                    import uuid
                    uuid.UUID(str(value).strip())
                except (ValueError, TypeError):
                    return f"Column '{column_name}' must be a valid UUID"
            
            # JSON type
            elif data_type_lower in ['json', 'jsonb']:
                try:
                    import json
                    json.loads(str(value).strip())
                except (ValueError, TypeError):
                    return f"Column '{column_name}' must be valid JSON"
            
            # Text types (character varying, text, etc.) - no specific validation needed
            # as they can accept any string
            
            return None  # No validation error
            
        except Exception as e:
            return f"Error validating column '{column_name}': {str(e)}"


class DeleteTableRecordView(APIView):
    """
    API view to delete a record from a specific table in the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def delete(self, request):
        try:
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            record_id = request.data.get('record_id')
            
            if not table_name or not record_id:
                return Response(
                    {"error": "Table name and record ID are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # First, get the primary key column name
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu 
                    ON tc.constraint_name = kcu.constraint_name
                    WHERE tc.table_name = %s 
                    AND tc.constraint_type = 'PRIMARY KEY'
                    AND tc.table_schema = %s
                """, [table_name, schema])
                
                pk_result = cursor.fetchone()
                if not pk_result:
                    return Response(
                        {"error": f"No primary key found for table '{table_name}'."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                pk_column = pk_result[0]
                
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'
                
                # Build DELETE query
                delete_query = f"""
                    DELETE FROM {table_reference} 
                    WHERE "{pk_column}" = %s
                """
                
                cursor.execute(delete_query, [record_id])
                
                if cursor.rowcount == 0:
                    return Response(
                        {"error": f"No record found with ID '{record_id}' in table '{table_name}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Log the record deletion activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, 1, user.email, f"Record deleted from table '{table_name}' (ID: {record_id})")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")
            
            customer_connection.close()
            
            return Response({
                "message": f"Record deleted successfully from table '{table_name}'.",
                "table_name": table_name,
                "record_id": record_id
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error deleting record: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


#updating user added table structure
class UpdateTableStructureView(APIView):
    """
    API view to update table structure by modifying column data types.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            schema = request.data.get('schema')
            table_name = request.data.get('table_name')
            column_changes = request.data.get('column_changes', {})
            column_renames = request.data.get('column_renames', {})
            columns_to_add = request.data.get('columns_to_add', [])
            columns_to_delete = request.data.get('columns_to_delete', [])
            
            if not schema or not table_name:
                return Response(
                    {"error": "Schema and table name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if at least one operation is requested
            if not column_changes and not column_renames and not columns_to_add and not columns_to_delete:
                return Response(
                    {"error": "At least one operation (column changes, renames, additions, or deletions) is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Check if table exists
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = %s 
                        AND table_name = %s
                    )
                """, [schema, table_name])
                
                if not cursor.fetchone()[0]:
                    return Response(
                        {"error": f"Table '{table_name}' does not exist."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Get current column information
                cursor.execute("""
                    SELECT column_name, data_type, character_maximum_length
                    FROM information_schema.columns
                    WHERE table_schema = %s 
                    AND table_name = %s
                    ORDER BY ordinal_position
                """, [schema, table_name])
                
                current_columns = cursor.fetchall()
                column_info = {col[0]: {'type': col[1], 'length': col[2]} for col in current_columns}
                
                # Process column operations
                updated_columns = []
                added_columns = []
                deleted_columns = []
                renamed_columns = []
                data_loss_warnings = []
                final_table_name = table_name
                
                # 1. Delete columns first (to avoid dependency issues)
                for column_name in columns_to_delete:
                    # Prevent deletion of system columns
                    if column_name in ['__id', '__active']:
                        data_loss_warnings.append(f"Cannot delete system column '{column_name}'.")
                        continue
                    
                    if column_name not in column_info:
                        data_loss_warnings.append(f"Column '{column_name}' does not exist in table '{table_name}'.")
                        continue
                    
                    # Check if column is part of primary key
                    cursor.execute("""
                        SELECT column_name 
                        FROM information_schema.table_constraints tc
                        JOIN information_schema.key_column_usage kcu 
                        ON tc.constraint_name = kcu.constraint_name
                        WHERE tc.table_name = %s 
                        AND tc.constraint_type = 'PRIMARY KEY'
                        AND tc.table_schema = %s
                        AND kcu.column_name = %s
                    """, [table_name, schema, column_name])
                    
                    if cursor.fetchone():
                        data_loss_warnings.append(f"Cannot delete column '{column_name}' as it is part of the primary key.")
                        continue
                    
                    try:
                        drop_query = f'ALTER TABLE "{schema}"."{final_table_name}" DROP COLUMN "{column_name}"'
                        cursor.execute(drop_query)
                        deleted_columns.append(column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to delete column '{column_name}': {str(e)}")
                
                # 2. Rename existing columns
                for old_column_name, new_column_name in column_renames.items():
                    if old_column_name not in column_info:
                        data_loss_warnings.append(f"Column '{old_column_name}' does not exist in table '{final_table_name}'.")
                        continue
                    
                    # Clean and validate new column name
                    new_column_name = new_column_name.strip().replace(' ', '_')
                    if not new_column_name.replace('_', '').isalnum():
                        data_loss_warnings.append(f"Invalid column name '{new_column_name}'. Only alphanumeric characters and underscores are allowed.")
                        continue
                    
                    # Check if new column name already exists
                    if new_column_name in column_info:
                        data_loss_warnings.append(f"Column '{new_column_name}' already exists in table '{final_table_name}'.")
                        continue
                    
                    try:
                        rename_query = f'ALTER TABLE "{schema}"."{final_table_name}" RENAME COLUMN "{old_column_name}" TO "{new_column_name}"'
                        cursor.execute(rename_query)
                        renamed_columns.append(f"{old_column_name} -> {new_column_name}")
                        
                        # Update column_info for subsequent operations
                        column_info[new_column_name] = column_info.pop(old_column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to rename column '{old_column_name}' to '{new_column_name}': {str(e)}")
                
                # 3. Add new columns
                for column_data in columns_to_add:
                    column_name = column_data.get('name', '').strip().replace(' ', '_')
                    column_type = column_data.get('type', 'VARCHAR(255)').strip()
                    
                    if not column_name:
                        data_loss_warnings.append("Column name cannot be empty.")
                        continue
                    
                    # Validate column name (basic validation)
                    if not column_name.replace('_', '').isalnum():
                        data_loss_warnings.append(f"Invalid column name '{column_name}'. Only alphanumeric characters and underscores are allowed.")
                        continue
                    
                    # Check if column already exists
                    if column_name in column_info:
                        data_loss_warnings.append(f"Column '{column_name}' already exists in table '{final_table_name}'.")
                        continue
                    
                    try:
                        add_query = f'ALTER TABLE "{schema}"."{final_table_name}" ADD COLUMN "{column_name}" {column_type}'
                        cursor.execute(add_query)
                        added_columns.append(column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to add column '{column_name}': {str(e)}")
                
                # 4. Update existing columns
                for column_name, new_type in column_changes.items():
                    if column_name not in column_info:
                        data_loss_warnings.append(f"Column '{column_name}' does not exist in table '{final_table_name}'.")
                        continue
                    
                    # Skip if the type hasn't changed
                    current_type = column_info[column_name]['type']
                    if current_type.upper() == new_type.upper():
                        continue
                    
                    # Check for potentially data-losing conversions
                    current_type_upper = current_type.upper()
                    new_type_upper = new_type.upper()
                    
                    # Check for decimal to integer conversions
                    if (current_type_upper in ['REAL', 'DOUBLE PRECISION', 'NUMERIC', 'DECIMAL'] and 
                        new_type_upper in ['INTEGER', 'BIGINT', 'SMALLINT', 'SERIAL', 'BIGSERIAL']):
                        data_loss_warnings.append(f"Converting '{column_name}' from {current_type} to {new_type} will truncate decimal values and cause data loss.")
                        continue
                    
                    # Check for text to numeric conversions
                    if (current_type_upper in ['TEXT', 'VARCHAR', 'CHARACTER VARYING'] and 
                        new_type_upper in ['INTEGER', 'BIGINT', 'SMALLINT', 'REAL', 'NUMERIC', 'DECIMAL']):
                        data_loss_warnings.append(f"Converting '{column_name}' from {current_type} to {new_type} may fail if text contains non-numeric values.")
                        continue
                    
                    # Check for numeric to text conversions (usually safe but warn)
                    if (current_type_upper in ['INTEGER', 'BIGINT', 'SMALLINT', 'REAL', 'NUMERIC', 'DECIMAL'] and 
                        new_type_upper in ['TEXT', 'VARCHAR', 'CHARACTER VARYING']):
                        data_loss_warnings.append(f"Converting '{column_name}' from {current_type} to {new_type} will convert numbers to text format.")
                        continue
                    
                    # Build ALTER COLUMN statement
                    alter_query = f'ALTER TABLE "{schema}"."{final_table_name}" ALTER COLUMN "{column_name}" TYPE {new_type}'
                    
                    try:
                        cursor.execute(alter_query)
                        updated_columns.append(column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to update column '{column_name}' to type '{new_type}': {str(e)}")
            
            customer_connection.close()
            
            # Check if any operations were successful
            total_operations = len(updated_columns) + len(added_columns) + len(deleted_columns) + len(renamed_columns)
            
            if total_operations == 0 and not data_loss_warnings:
                return Response({
                    "message": "No changes were made to the table structure.",
                    "table_name": final_table_name
                }, status=status.HTTP_200_OK)
            
            response_data = {
                "table_name": final_table_name,
                "updated_columns": updated_columns,
                "added_columns": added_columns,
                "deleted_columns": deleted_columns,
                "renamed_columns": renamed_columns
            }
            
            if data_loss_warnings:
                response_data["warnings"] = data_loss_warnings
                if total_operations > 0:
                    operations = []
                    if renamed_columns:
                        operations.append(f"{len(renamed_columns)} column(s) renamed")
                    if updated_columns:
                        operations.append(f"{len(updated_columns)} column(s) modified")
                    if added_columns:
                        operations.append(f"{len(added_columns)} column(s) added")
                    if deleted_columns:
                        operations.append(f"{len(deleted_columns)} column(s) deleted")
                    
                    response_data["message"] = f"Table structure updated successfully. {', '.join(operations)}. Some operations were skipped due to errors or potential data loss."
                else:
                    response_data["message"] = "No changes were made due to errors or potential data loss in the requested operations."
                    return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            else:
                operations = []
                if renamed_columns:
                    operations.append(f"{len(renamed_columns)} column(s) renamed")
                if updated_columns:
                    operations.append(f"{len(updated_columns)} column(s) modified")
                if added_columns:
                    operations.append(f"{len(added_columns)} column(s) added")
                if deleted_columns:
                    operations.append(f"{len(deleted_columns)} column(s) deleted")
                
                response_data["message"] = f"Table structure updated successfully. {', '.join(operations)}."
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error updating table structure: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


#deleting user added tables
class DeleteTableView(APIView):
    """
    API view to delete a table from the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def delete(self, request):
        try:
            schema = request.data.get('schema')
            table_name = request.data.get('table_name')
            
            if not schema or not table_name:
                return Response(
                    {"error": "Schema and table name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Check if table exists
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = %s 
                        AND table_name = %s
                    )
                """, [schema, table_name])
                
                if not cursor.fetchone()[0]:
                    return Response(
                        {"error": f"Table '{table_name}' does not exist."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Get record count before deletion
                cursor.execute(f'SELECT COUNT(*) FROM "{schema}"."{table_name}";')
                record_count = cursor.fetchone()[0]
                
                # Drop the table
                drop_query = f'DROP TABLE "{schema}"."{table_name}";'
                cursor.execute(drop_query)
                
                # Log the table deletion activity
                try:
                    insert_activity_log_sql = f'''
                        INSERT INTO "{schema}"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, record_count, user.email, f"Table '{table_name}' deleted with {record_count} records")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")
            
            customer_connection.close()
            
            return Response({
                "message": f"Table '{table_name}' deleted successfully. {record_count} records were removed.",
                "table_name": table_name,
                "deleted_records": record_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Error deleting table: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CreateTableWithoutRecordsView(APIView):
    """
    API view to create a directory for a table in the user's customer database.
    """

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self,request):

        table_name = request.data.get('table_name')
        columns = request.data.get('column_names')
        scope = request.data.get('scope')
        schema = request.data.get('schema')

    

        if not table_name or not columns or not schema:
            return Response(
                {"error": "Table name and column names and schema are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        #validate table name format 
        clean_table_name = table_name.strip().lower().replace(" ",'_')
        if not clean_table_name.replace("_","").isalnum():
            return Response(
                {"error": "Table name must contain only letters, numbers, and underscores."},
            )

        if not clean_table_name[0].isalpha() or clean_table_name[0] == "_":
            return Response(
                {"error": "Table name must start with a letter or underscore."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Add __ suffix to table name
        clean_table_name = clean_table_name + "__"

        column_names = []
        for col in columns:
            col_name = col.get("column_name")
            if col_name.lower() in ['__id', 'is_active']: # reserved fields
                return Response(
                    {"error": f"Column name {col_name} is reserved."},
                )
            data_type = col.get('data_type', '')

            if not col_name or not data_type:
                return Response(
                    {"error": "Column name and data type are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if col_name.lower() in column_names:
                return Response(
                    {"error": f"duplicate column name: {col_name}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            column_names.append(col_name.lower())


        try:
            user = request.user
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            customer = user.cust_id  # already a Customer instance
        except User.DoesNotExist:
            return Response(
                {"error": "User not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Connect to customer's database
        customer_connection = psycopg2.connect(
            host=settings.DATABASES['default']['HOST'],
            port=settings.DATABASES['default']['PORT'],
            database=customer.cust_db,
            user=settings.DATABASES['default']['USER'],
            password=settings.DATABASES['default']['PASSWORD']
        )
        customer_connection.autocommit = True
        
        # Determine target schema based on scope
        target_schema = 'GENERAL' if scope == 'global' else schema

        with customer_connection.cursor() as cursor:
            # Set search path to target schema
            cursor.execute(f'SET search_path TO "{target_schema}";')
            
            # Check if table exists
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_schema = %s 
                    AND table_name = %s
                )
            """, [target_schema, clean_table_name])

            if cursor.fetchone()[0]:
                return Response(
                    {"error": f"Table '{table_name}' already exists."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            #create the column definition - start with auto-increment ID as primary key
            column_definitions = ['__id SERIAL PRIMARY KEY']
            for col in columns:
                col_name = col.get("column_name")
                data_type = col.get('data_type', '')
                is_nullable = col.get('is_nullable', True)
                column_definitions.append(f'"{col_name}" {data_type} {("NOT NULL" if not is_nullable else "")}')
            
            create_table_sql = f'CREATE TABLE IF NOT EXISTS "{target_schema}"."{clean_table_name}" ({", ".join(column_definitions)});'
            cursor.execute(create_table_sql)

            #add is_active field to the table
            cursor.execute(f'''
                ALTER TABLE "{target_schema}"."{clean_table_name}" ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE;
            ''')

            #create the activity_log table if doesnt exist
            create_log_table_sql = f'''
                CREATE TABLE IF NOT EXISTS "{target_schema}"."activity_log" (
                    table_name VARCHAR(100) NOT NULL,
                    created_by VARCHAR(100),
                    created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    records_count INTEGER NOT NULL,
                    modified_by VARCHAR(100),
                    modified_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    description TEXT
                );
            '''
            cursor.execute(create_log_table_sql)

            #insert the log entry
            insert_log_sql = f'''
                INSERT INTO "{target_schema}"."activity_log" (table_name, created_by, records_count, modified_by, description)
                VALUES (%s, %s, %s, %s, %s)
            '''
            cursor.execute(
                insert_log_sql,
                (clean_table_name, user.email, 0, None, f"Table '{clean_table_name}' created with {len(columns) + 1} columns (including auto-increment ID)")
            )

        customer_connection.close()

        return Response(
            {"message": f"Table '{clean_table_name}' created successfully in schema '{target_schema}' with {len(columns) + 1} columns (including auto-increment ID).",
            "schema":schema},
            status=status.HTTP_200_OK
        )


class ImportDataFromHanaView(APIView):
    def post(self, request):
        user_id = request.data.get('user_id')
        print(f"User ID received: {user_id}")
        
        if not user_id:
            return Response(
                {"message": "User ID is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response(
            {"message": f"Import data from HANA for user ID: {user_id}"},
            status=status.HTTP_200_OK
        )


class DownloadTableDataView(APIView):
    """
    API endpoint to download table data in multiple formats (CSV, Excel, TSV) with filtering support
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            filters = request.data.get('filters', [])
            sort_column = request.data.get('sort_column', '')
            sort_direction = request.data.get('sort_direction', 'asc')
            file_format = request.data.get('format', 'csv').lower()  # Default to CSV
            selected_columns = request.data.get('selected_columns', None)  # Get selected columns
            
            if not table_name:
                return Response(
                    {"error": "table_name is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate file format
            if file_format not in ['csv', 'excel', 'xlsx', 'tsv']:
                return Response(
                    {"error": "Invalid format. Supported formats: csv, excel, xlsx, tsv"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
                
                # Update user's file format preference if format is provided
                if file_format in ['csv', 'excel', 'xlsx', 'tsv']:
                    # Normalize excel/xlsx to 'excel'
                    normalized_format = 'excel' if file_format in ['excel', 'xlsx'] else file_format
                    user.file_format = normalized_format
                    user.save(update_fields=['file_format'])
                
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found"},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                
                return Response(
                    {"error": "Customer not found"},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Connect to customer database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Get table columns (excluding __active and __id as they will be added automatically)
                cursor.execute("""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_schema = %s 
                    AND table_name = %s 
                    AND column_name NOT IN ('__active', '__id')
                    ORDER BY ordinal_position;
                """, (schema, table_name))
                
                table_columns = cursor.fetchall()
                
                if not table_columns:
                    customer_connection.close()
                    return Response(
                        {"error": f"Table '{table_name}' not found in schema '{schema}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Filter columns based on selected_columns if provided
                if selected_columns and isinstance(selected_columns, list) and len(selected_columns) > 0:
                    # Create a dictionary for quick lookup
                    columns_dict = {col[0]: col for col in table_columns}
                    
                    # Filter and reorder table_columns to match selected_columns order
                    filtered_columns = []
                    for col_name in selected_columns:
                        if col_name in columns_dict:
                            filtered_columns.append(columns_dict[col_name])
                    
                    table_columns = filtered_columns
                    
                    # If no columns match after filtering, return error
                    if not table_columns:
                        customer_connection.close()
                        return Response(
                            {"error": "None of the selected columns exist in the table."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # Build WHERE clause for filters
                where_conditions = []
                filter_params = []
                
                for filter_item in filters:
                    column = filter_item.get('column')
                    operator = filter_item.get('operator')
                    value = filter_item.get('value')
                    
                    if column and operator and value is not None:
                        if operator == 'ILIKE':
                            where_conditions.append(f'"{column}" ILIKE %s')
                            filter_params.append(f'%{value}%')
                        elif operator == 'IN':
                            # Handle both list and comma-separated string
                            if isinstance(value, list) and value:
                                placeholders = ','.join(['%s'] * len(value))
                                where_conditions.append(f'"{column}" IN ({placeholders})')
                                filter_params.extend(value)
                            elif isinstance(value, str) and value:
                                values_list = [v.strip() for v in value.split(',') if v.strip()]
                                if values_list:
                                    placeholders = ','.join(['%s'] * len(values_list))
                                    where_conditions.append(f'"{column}" IN ({placeholders})')
                                    filter_params.extend(values_list)
                        elif operator == 'NOT IN':
                            # Handle both list and comma-separated string
                            if isinstance(value, list) and value:
                                placeholders = ','.join(['%s'] * len(value))
                                where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                                filter_params.extend(value)
                            elif isinstance(value, str) and value:
                                values_list = [v.strip() for v in value.split(',') if v.strip()]
                                if values_list:
                                    placeholders = ','.join(['%s'] * len(values_list))
                                    where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                                    filter_params.extend(values_list)
                        elif operator == 'IS NULL':
                            where_conditions.append(f'"{column}" IS NULL')
                        elif operator == 'IS NOT NULL':
                            where_conditions.append(f'"{column}" IS NOT NULL')
                        elif operator == 'MISMATCH':
                            # Handle MISMATCH operator - value contains JSON string with mismatch pairs
                            try:
                                import json
                                mismatch_data = json.loads(value) if isinstance(value, str) else value
                                if isinstance(mismatch_data, list) and len(mismatch_data) > 0:
                                    # Get the first item to determine related column name
                                    first_item = mismatch_data[0]
                                    related_column = None
                                    for key in first_item.keys():
                                        if key != column:
                                            related_column = key
                                            break
                                    
                                    if related_column:
                                        # Build OR conditions for each mismatch pair
                                        mismatch_conditions = []
                                        for mismatch_pair in mismatch_data:
                                            mismatch_conditions.append(
                                                f'("{column}" = %s AND "{related_column}" = %s)'
                                            )
                                            filter_params.append(mismatch_pair.get(column))
                                            filter_params.append(mismatch_pair.get(related_column))
                                        
                                        # Join with OR
                                        where_conditions.append(f'({" OR ".join(mismatch_conditions)})')
                            except (json.JSONDecodeError, AttributeError, KeyError) as e:
                                # If JSON parsing fails, skip this filter
                                continue
                        else:
                            where_conditions.append(f'"{column}" {operator} %s')
                            filter_params.append(value)
                
                # Build ORDER BY clause
                order_by = ''
                if sort_column:
                    order_direction = 'ASC' if sort_direction.lower() == 'asc' else 'DESC'
                    order_by = f'ORDER BY "{sort_column}" {order_direction}'
                
                # Build the complete query
                where_clause = 'WHERE ' + ' AND '.join(where_conditions) if where_conditions else ''
                
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'
                
                query = f'''
                    SELECT {', '.join([f'"{col[0]}"' for col in table_columns])}
                    FROM {table_reference}
                    {where_clause}
                    {order_by}
                '''
                
                # Execute query
                cursor.execute(query, filter_params)
                rows = cursor.fetchall()
                
                # Get headers
                headers = [col[0] for col in table_columns]
                
                # Convert rows to list of dictionaries for date formatting
                data = []
                for row in rows:
                    row_dict = {}
                    for i, value in enumerate(row):
                        row_dict[headers[i]] = value
                    data.append(row_dict)
                
                # Format date/timestamp columns based on user's preferred date format
                user_strftime_format = convert_user_date_format_to_strftime(user.date_format)
                data = format_date_columns(data, table_columns, user_strftime_format)
                
                # Convert back to rows format
                formatted_rows = []
                for row_dict in data:
                    formatted_rows.append([row_dict.get(header) for header in headers])
                
                # Generate file content based on format
                if file_format in ['csv']:
                    file_content, content_type, file_extension = self._generate_csv(headers, formatted_rows)
                elif file_format in ['excel', 'xlsx']:
                    file_content, content_type, file_extension = self._generate_excel(headers, formatted_rows)
                elif file_format == 'tsv':
                    file_content, content_type, file_extension = self._generate_tsv(headers, formatted_rows)
            
            customer_connection.close()
            
            # Create HTTP response with appropriate content
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{table_name}_data.{file_extension}"'
            
            return response
            
        except Exception as e:
            return Response(
                {"error": f"Failed to download data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generate_csv(self, headers, rows):
        """Generate CSV content"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header row
        writer.writerow(headers)
        
        # Write data rows
        for row in rows:
            # Convert None values to 'None' string for CSV
            csv_row = [str(val) if val is not None else 'None' for val in row]
            writer.writerow(csv_row)
        
        content = output.getvalue()
        output.close()
        
        return content, 'text/csv', 'csv'
    
    def _generate_tsv(self, headers, rows):
        """Generate TSV (Tab-Separated Values) content"""
        import io
        
        output = io.StringIO()
        
        # Write header row
        output.write('\t'.join(headers) + '\n')
        
        # Write data rows
        for row in rows:
            # Convert None values to 'None' string and join with tabs
            tsv_row = [str(val) if val is not None else 'None' for val in row]
            output.write('\t'.join(tsv_row) + '\n')
        
        content = output.getvalue()
        output.close()
        
        return content, 'text/tab-separated-values', 'tsv'
    
    def _generate_excel(self, headers, rows):
        """Generate Excel content"""
        try:
            import openpyxl
            from io import BytesIO
        except ImportError:
            # Fallback to basic CSV if openpyxl is not available
            return self._generate_csv(headers, rows)
        
        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)
        
        # Write data rows
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                # Convert None values to 'None' string for Excel
                cell_value = 'None' if value is None else value
                worksheet.cell(row=row_num, column=col_num, value=cell_value)
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        content = output.getvalue()
        output.close()
        
        return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx'


class LoginView(APIView):
    """
    API endpoint for user login with JWT authentication.
    """
    authentication_classes = []  # No authentication required for login
    permission_classes = []  # No permissions required for login
    
    def post(self, request):
        """
        Handle user login and return JWT tokens.
        """
        email = request.data.get('email').lower()
        print(f"Email received: {email}")
        password = request.data.get('password')
        
        if not email or not password:
            return Response({
                'error': 'Please provide both email and password'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Authenticate user
        user = authenticate(request, username=email, password=password)

        if not user:
            return Response({
                'error': 'Invalid email or password'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        if user is not None:
            if user.is_active:
                # Update last_login field
                user.last_login = timezone.now()
                user.save(update_fields=['last_login'])
                
                # Generate JWT tokens
                refresh = RefreshToken.for_user(user)
                
                customer = user.cust_id
                
                # Create response
                response = Response({
                    'message': 'Login successful',
                    'user': {
                        'email': user.email,
                        'first_name': user.first_name,
                        'last_name': user.last_name,
                        'username': user.username,
                    }
                }, status=status.HTTP_200_OK)
                
                # Set HTTP-only cookies for tokens
                response.set_cookie(
                    key='access_token',
                    value=str(refresh.access_token),
                    httponly=True,
                    secure=False,  # Set to True in production with HTTPS
                    samesite='Lax',
                    max_age=3600  # 1 hour (matches ACCESS_TOKEN_LIFETIME)
                )
                
                response.set_cookie(
                    key='refresh_token',
                    value=str(refresh),
                    httponly=True,
                    secure=False,  # Set to True in production with HTTPS
                    samesite='Lax',
                    max_age=86400  # 1 day (matches REFRESH_TOKEN_LIFETIME)
                )
                
                return response
            else:
                return Response({
                    'error': 'User account is disabled'
                }, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({
                'error': 'Invalid email or password'
            }, status=status.HTTP_401_UNAUTHORIZED)


class LogoutView(APIView):
    """
    API endpoint for user logout. 
    Clears JWT tokens stored in cookies and Django session.
    """    
    authentication_classes = []  # No authentication required for logout
    permission_classes = []  # No permissions required for logout
    
    def get(self, request):
        """
        Handle user logout by clearing JWT token cookies and Django session.
        """
        try:
            
            # Flush the Django session if it exists
            if hasattr(request, 'session'):
                request.session.flush()
                print("Django session flushed")
            
            # Create response
            response = Response({
                'message': 'Logout successful'
            }, status=status.HTTP_200_OK)
            
            # Delete the access_token cookie
            # IMPORTANT: All parameters must match those used when setting the cookie
            response.delete_cookie(
                key='access_token',
                path='/',
                samesite='Lax'
            )
            
            # Delete the refresh_token cookie
            # IMPORTANT: All parameters must match those used when setting the cookie
            response.delete_cookie(
                key='refresh_token',
                path='/',
                samesite='Lax'
            )
            
            # Also delete the sessionid cookie (Django's default session cookie)
            response.delete_cookie(
                key='sessionid',
                path='/',
                samesite='Lax'
            )
            
            # Delete CSRF cookie
            response.delete_cookie(
                key='csrftoken',
                path='/',
                samesite='Lax'
            )

            return response
            
        except Exception as e:
            return Response({
                'error': f'Logout failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    
class RefreshTokenView(APIView):
    """
    API endpoint to refresh access token using the refresh token stored in an HttpOnly cookie.
    """
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        try:
            refresh_token = request.COOKIES.get('refresh_token')
            if not refresh_token:
                return Response({'error': 'Refresh token missing'}, status=status.HTTP_401_UNAUTHORIZED)

            serializer = TokenRefreshSerializer(data={'refresh': refresh_token})
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data

            response = Response({'message': 'Token refreshed successfully'}, status=status.HTTP_200_OK)

            # Set new access token
            response.set_cookie(
                key='access_token',
                value=data.get('access'),
                httponly=True,
                secure=False,  # Set to True in production with HTTPS
                samesite='Lax',
                max_age=3600
            )

            # If rotation is enabled, a new refresh may be returned; update cookie
            new_refresh = data.get('refresh')
            if new_refresh:
                response.set_cookie(
                    key='refresh_token',
                    value=new_refresh,
                    httponly=True,
                    secure=False,
                    samesite='Lax',
                    max_age=86400
                )

            return response
        except Exception as e:
            return Response({'error': f'Invalid or expired refresh token: {str(e)}'}, status=status.HTTP_401_UNAUTHORIZED)

    
class TruncateTableView(APIView):
    """
    API endpoint to truncate a table.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        table_name = request.data.get('table_name')
        schema = request.data.get('schema')
        filters = request.data.get('filters', [])
        
        if not table_name:
            return Response(
                {"error": "table_name is required"},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            user = request.user
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            customer = user.cust_id
            cust_db = customer.cust_db
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND)
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"},
                status=status.HTTP_404_NOT_FOUND)

        try:
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
        
            with customer_connection.cursor() as cursor:
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'

                # Get count before truncation/deletion
                cursor.execute(f'SELECT COUNT(*) FROM {table_reference}')
                records_count = cursor.fetchone()[0]

                if filters:
                    where_conditions = []
                    filter_params = []
                    for filter in filters:
                        column = filter.get('column')
                        operator = filter.get('operator')
                        value = filter.get('value')

                        if operator == 'ILIKE':
                            where_conditions.append(f'"{column}" ILIKE %s')
                            filter_params.append(f'%{value}%')
                        elif operator == 'IN':
                            value_list = [v.strip() for v in value.split(',') if v.strip()]
                            placeholders = ','.join(['%s'] * len(value_list))
                            where_conditions.append(f'"{column}" IN ({placeholders})')
                            filter_params.extend(value_list)
                        elif operator == 'NOT IN':
                            value_list = [v.strip() for v in value.split(',') if v.strip()]
                            placeholders = ','.join(['%s'] * len(value_list))
                            where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                            filter_params.extend(value_list)
                        elif operator == 'IS NULL':
                            where_conditions.append(f'"{column}" IS NULL')
                        elif operator == 'IS NOT NULL':
                            where_conditions.append(f'"{column}" IS NOT NULL')
                        
                        else:   
                            where_conditions.append(f'"{column}" {operator} %s')
                            filter_params.append(value)

                    cursor.execute(f'SELECT COUNT(*) FROM {table_reference} WHERE {" AND ".join(where_conditions)}', filter_params)
                    deleted_count = cursor.fetchone()[0]
                    delete_query = f'DELETE FROM {table_reference} WHERE {" AND ".join(where_conditions)}'
                    cursor.execute(delete_query, filter_params)
                        


                else:
                    cursor.execute(f'TRUNCATE TABLE {table_reference}')
                    deleted_count = records_count
                
                # Log the truncate/delete activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    if filters:
                        description = f"Deleted {deleted_count} record(s) from table '{table_name}' with filters"
                    else:
                        description = f"Table '{table_name}' truncated with {deleted_count} record(s) removed"
                    
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, deleted_count, user.email, description)
                    )
                except Exception as log_error:
                    print(f"Error logging truncate activity: {log_error}")
            
            customer_connection.close()

            
            
            return Response({"message": "Table truncated successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"Failed to truncate table: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CreateUserView(APIView):
    """
    API endpoint for creating users.
    """

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """
        Create a new user.
        """
        serializer = UserSerializer(data=request.data)

        
        if serializer.is_valid():
            try:
                serializer.validated_data['email'] = serializer.validated_data['email'].lower()
                serializer.validated_data['cust_id'] = request.user.cust_id
                serializer.validated_data['created_by'] = request.user.email
                serializer.save()
            except Exception as e:
                return Response(
                    {"error": f"Error creating user: {str(e)}"}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UserListView(APIView):

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self,request):

        user = request.user
        customer = user.cust_id
        if not customer:
            return Response(
                {"error":"user is not associated with any customer"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            company_users = User.objects.filter(cust_id=customer).exclude(email=user.email)
            serializer = UserSerializer(company_users, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"Error retrieving users: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class UserUpdateView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, user_id):
        user = User.objects.get(id=user_id)
        serializer = UserSerializer(user)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def put(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            serializer = UserSerializer(user, data=request.data, partial=True)
            
            if serializer.is_valid():
                serializer.validated_data['modified_by'] = request.user.email
                serializer.save()
                return Response({
                    'message': 'User updated successfully',
                    'data': serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'error': 'Validation failed',
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except User.DoesNotExist:
            return Response({
                'error': 'User not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'error': f'Error updating user: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UserDeleteView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            user.delete()
            return Response({
                'message': 'User deleted successfully'
            }, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({
                'error': 'User not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'error': f'Error deleting user: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes

from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail

class UserPasswordResetView(APIView):
    """
    API endpoint for resetting user password.
    """
    def post(self, request):
        email = request.data.get('email').lower()
        if not email:
            return Response({
                'error': 'Email is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({
                'error': 'User with this email address does not exist'
            }, status=status.HTTP_404_NOT_FOUND)

        uid = urlsafe_base64_encode(force_bytes(user.email))
        token = default_token_generator.make_token(user)

        FRONTEND_URL = 'http://localhost:8000/api'

        reset_url = f"{FRONTEND_URL}/reset-password-confirm/{uid}/{token}"

        # Send password reset email
        try:
            send_mail(
                subject='Password Reset Request',
                message=f'Hello,\n\nYou requested to reset your password. Click the link below to reset your password:\n\n{reset_url}\n\nIf you did not request this, please ignore this email.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
        except Exception as e:
            return Response(
                {"error": "Failed to send email. Please try again later." + str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        return Response(
            {"message": "Password reset link has been sent to your email."}, 
            status=status.HTTP_200_OK
        )


class UserPasswordResetConfirmView(APIView):
    """
    API endpoint for confirming password reset with token.
    """
    
    def post(self, request):
        """
        Confirm password reset and set new password.
        """
        uidb64 = request.data.get('uid')
        token = request.data.get('token')
        new_password = request.data.get('new_password')
        
        # Validate required fields
        if not all([uidb64, token, new_password]):
            return Response(
                {"error": "uid, token, and new_password are required."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Decode the email from base64
            email = urlsafe_base64_decode(uidb64).decode()
            user = User.objects.get(email=email)
        except (User.DoesNotExist, ValueError, TypeError, UnicodeDecodeError):
            return Response(
                {"error": "Invalid reset link."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verify token
        if not default_token_generator.check_token(user, token):
            return Response(
                {"error": "Invalid or expired reset link."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Set new password
        user.set_password(new_password)
        user.save()
        
        return Response(
            {"message": "Password has been reset successfully."}, 
            status=status.HTTP_200_OK
        )


class ProjectsListView(APIView):
    """
    API endpoint for listing projects.
    """

    # authentication_classes = [JWTCookieAuthentication]
    # permission_classes = [IsAuthenticated]


    def get(self, request):

        user = request.user
        customer = user.cust_id
        if not customer:
            return Response(
                {"error":"user is not associated with any customer"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        connection = psycopg2.connect(
            host=settings.DATABASES['default']['HOST'],
            port=settings.DATABASES['default']['PORT'],
            user=settings.DATABASES['default']['USER'],
            password=settings.DATABASES['default']['PASSWORD'],
            database=customer.cust_db
        )
        cursor = connection.cursor()
        cursor.execute('''SELECT "ID","IDENT","DESCR","IS_ACTIVE","REF_CLIENT" FROM "GENERAL"."PROJECT"''')
        projects_raw = cursor.fetchall()
        cursor.close()
        connection.close()
        
        # Format projects data similar to source data
        projects = []
        for project in projects_raw:
            project_data = {
                'id': project[0],
                'ident': project[1],
                'descr': project[2],
                'is_active': project[3],
                'ref_client': project[4]
            }
            projects.append(project_data)
        
        return Response(projects, status=status.HTTP_200_OK)


class FilterColumnValuesView(APIView):
    """
    API endpoint for filtering column values.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]



# Helper functions for column statistics
def get_column_info(column_name):
    """
    Get column information from column_categories.json
    """
    with open('column_categories.json', 'r') as f:
        column_categories = json.load(f)
        for category, value in column_categories.items():
            if isinstance(value,list) and column_name in value:
                return category
            elif isinstance(value,dict) and "columns" in value and column_name in value["columns"]:
                return category
    return None
 

 # Helper function for column statistics
def get_category_columns(column_name):
    """
    Get category columns from column_categories.json
    """
    with open('column_categories.json', 'r') as f:
        column_categories = json.load(f)
        return column_categories.get(column_name, None)

# hepler function for column statistics
def check_columns_in_table(cursor, schema, table_name, category_columns):
    """
    Check if any columns from category_columns exist in the table
    """
    if not category_columns:
        return []
    
    # Create placeholders for the IN clause
    placeholders = ','.join(['%s'] * len(category_columns))
    
    query = f"""
        SELECT column_name 
        FROM information_schema.columns 
        WHERE table_schema = %s 
        AND table_name = %s 
        AND column_name IN ({placeholders})
    """
    
    cursor.execute(query, [schema, table_name] + category_columns)
    existing_columns = [row[0] for row in cursor.fetchall()]

    if len(existing_columns) > 0:
        return existing_columns[0]
    return None

class ColumnStatisticsView(APIView):
    """
    API endpoint for getting column statistics.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        column_name = request.data.get('column_name')
        table_name = request.data.get('table_name')
        schema = request.data.get('schema')
        # user_email = request.data.get('user')

        if not column_name or not table_name or not schema:
            return Response(
                {"error": "Column name, table name and schema are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        column_category = get_column_info(column_name)

        if not column_category:
            return Response(
                {"error": "Column category not found."},
                status=status.HTTP_400_BAD_REQUEST
            )
        user = request.user
        # user=User.objects.get(email=user_email)
        customer = user.cust_id
        if not customer:
            return Response(
                {"error": "User is not associated with any customer."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD'],
                database=customer.cust_db
            )
            with connection.cursor() as cursor:

                if column_category == 'numerical':
                    cursor.execute(f"""
                        SELECT 
                            MIN("{column_name}"), 
                            MAX("{column_name}"), 
                            COUNT(*),
                            COUNT(CASE WHEN "{column_name}" = 0 THEN 1 END) as zero_count,
                            COUNT(CASE WHEN "{column_name}" IS NULL THEN 1 END) as null_count
                        FROM "{schema}"."{table_name}"
                    """)
                    
                    result = cursor.fetchone()
                    min_val, max_val, total_count, zero_count, null_count = result
                    
                    return Response({
                        "column_name": column_name,
                        "column_category": column_category,
                        "statistics": {
                            "min_value": min_val,
                            "max_value": max_val,
                            "total_count": total_count,
                            "zero_count": zero_count,
                            "null_count": null_count
                        }
                    })

                elif column_category == 'string':
                    # Get basic statistics
                    cursor.execute(f"""
                        SELECT 
                            COUNT(*) as total_count,
                            COUNT(CASE WHEN "{column_name}" IS NULL THEN 1 END) as null_count
                        FROM "{schema}"."{table_name}"
                    """)
                    
                    result = cursor.fetchone()
                    total_count, null_count = result
                    
                    # Get distinct values and their counts
                    cursor.execute(f"""
                        SELECT 
                            "{column_name}" as value,
                            COUNT(*) as count
                        FROM "{schema}"."{table_name}"
                        WHERE "{column_name}" IS NOT NULL
                        GROUP BY "{column_name}"
                        ORDER BY "{column_name}" ASC
                    """)
                    
                    distinct_values = cursor.fetchall()
                    return Response({
                        "column_name": column_name,
                        "column_category": column_category,
                        "statistics": {
                            "total_count": total_count,
                            "null_count": null_count,
                            "distinct_values": [
                                {"value": row[0], "count": row[1]} 
                                for row in distinct_values
                            ]
                        }
                    })

                elif column_category == 'date':
                    # Load data using pandas for better date analysis
                    cursor.execute(f"""
                        SELECT "{column_name}"
                        FROM "{schema}"."{table_name}"
                    """)

                    
                    # Fetch all data and create DataFrame
                    data = cursor.fetchall()
                    df = pd.DataFrame(data, columns=[column_name])
                    
                    # Basic statistics
                    total_count = len(df)
                    null_count = df[column_name].isnull().sum()
                    non_null_count = df[column_name].notna().sum()
                    
                    # Count invalid dates and get min/max dates using pandas
                    non_null_data = df[column_name].dropna()
                    invalid_dates = []
                    invalid_date_count = 0
                    min_date = None
                    max_date = None
                    
                    if len(non_null_data) > 0:
                        # Convert to datetime with errors='coerce' - invalid dates become NaT
                        converted_dates = pd.to_datetime(non_null_data, errors='coerce')
                        
                        # Find invalid dates (where conversion resulted in NaT)
                        invalid_mask = converted_dates.isnull()
                        invalid_date_count = invalid_mask.sum()
                        
                        # Get valid dates (non-NaT)
                        valid_dates = converted_dates.dropna()
                        
                        if len(valid_dates) > 0:
                            min_date = str(valid_dates.min())
                            max_date = str(valid_dates.max())
                        
                    # Get the actual invalid date values for display
                    if invalid_date_count > 0:
                        invalid_dates = non_null_data[invalid_mask].tolist()
                        # Convert to string and remove duplicates
                        invalid_dates = list(set([str(date) for date in invalid_dates]))
                    
                    return Response({
                        "column_name": column_name,
                        "column_category": column_category,
                        "statistics": {
                            "total_count": int(total_count),
                            "null_count": int(null_count),
                            "non_null_count": int(non_null_count),
                            "min_date": min_date,
                            "max_date": max_date,
                            "invalid_date_count": int(invalid_date_count),
                            "invalid_dates": invalid_dates
                        }
                    })

                elif column_category == 'country_code':
                    # Load data using pandas for better country code analysis
                    cursor.execute(f"""
                        SELECT "{column_name}"
                        FROM "{schema}"."{table_name}"
                    """)
                    
                    # Fetch all data and create DataFrame
                    data = cursor.fetchall()
                    df = pd.DataFrame(data, columns=[column_name])
                    # Check if maximum character length in this column is more than three
                    max_char_len = df[column_name].dropna().astype(str).apply(len).max() if not df.empty else 0

                    if max_char_len > 3:
                        return Response(
                            {"error": "Maximum character length in this column is less than three."},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                      
                    # Basic statistics - null count and invalid country codes
                    null_count = df[column_name].isnull().sum()
                    
                    # Get distinct country codes with counts
                    distinct_countries = []
                    # Get distinct country codes and their counts
                    country_counts = df[column_name].dropna().astype(str).str.strip().str.upper().value_counts()
                    distinct_countries = [
                        {"value": code, "count": int(count)}
                        for code, count in country_counts.items()
                    ]
                    
                    # Validate country codes (ISO 3166-1 Alpha-3)
                    non_null_data = df[column_name].dropna()
                    invalid_country_codes = []
                    invalid_country_code_count = 0
                    
                    if len(non_null_data) > 0:
                        # Get all valid ISO 3166-1 Alpha-3 country codes
                        valid_country_codes = {country.alpha_2 for country in pycountry.countries} | {country.alpha_3 for country in pycountry.countries}

                        # Convert to string and strip whitespace, then to uppercase
                        non_null_data_str = non_null_data.astype(str).str.strip().str.upper()
                        
                        # Check length and validity
                        length_invalid_mask = non_null_data_str.str.len() > 3
                        not_in_valid_mask = ~non_null_data_str.isin(valid_country_codes)

                        # Combine masks using OR operator to find invalid entries
                        invalid_mask = length_invalid_mask | not_in_valid_mask

                        # Extract unique invalid country codes
                        invalid_country_codes = non_null_data_str[invalid_mask].unique().tolist()
                        invalid_country_code_count = len(invalid_country_codes)

                   
                     
                    return Response({
                        "column_name": column_name,
                        "column_category": column_category,
                        "statistics": {
                            "null_count": int(null_count),
                            "invalid_country_code_count": int(invalid_country_code_count),
                            "invalid_country_codes": invalid_country_codes,
                            "distinct_countries": distinct_countries
                        }
                        # "mismatched_country_codes": mismatched_country_codes 
                    })

                elif column_category == 'currency_code':
                    # Load data using pandas for better currency code analysis
                    cursor.execute(f"""
                        SELECT "{column_name}"
                        FROM "{schema}"."{table_name}"
                    """)
                    
                    # Fetch all data and create DataFrame
                    data = cursor.fetchall()
                    df = pd.DataFrame(data, columns=[column_name])
                    
                    # Basic statistics - null count
                    null_count = df[column_name].isnull().sum()
                    
                    # Get distinct currency codes and their counts
                    currency_counts = df[column_name].dropna().astype(str).str.strip().value_counts()
                    distinct_currencies = [
                        {"value": code, "count": int(count)}
                        for code, count in currency_counts.items()
                    ]
                    
                    # Validate currency codes
                    non_null_data = df[column_name].dropna()
                    invalid_currency_codes = []
                    invalid_currency_code_count = 0
                    
                    if len(non_null_data) > 0:
                        # Get all valid currency codes from pycountry
                        valid_currencies = {currency.alpha_3 for currency in pycountry.currencies}

                        
                        # Convert to string and strip whitespace, then to uppercase
                        non_null_data_str = non_null_data.astype(str).str.strip().str.upper()
                        
                        # Check validity of currency codes
                        not_in_valid_mask = ~non_null_data_str.isin(valid_currencies)
                        
                        # Extract unique invalid currency codes (original values as fetched from DB)
                        invalid_currency_codes = non_null_data[not_in_valid_mask].unique().tolist()
                        invalid_currency_code_count = len(invalid_currency_codes)
                        
    
                    return Response({
                        "column_name": column_name,
                        "column_category": column_category,
                        "statistics": {
                            "null_count": int(null_count),
                            "invalid_currency_code_count": int(invalid_currency_code_count),
                            "invalid_currency_codes": invalid_currency_codes,
                            "distinct_currencies": distinct_currencies
                        }
                    })
                elif column_category == 'region_category':
                    # First check if country_column_in_table exists
                    country_columns = get_category_columns('country_code')
                    country_column_in_table = None
                    if country_columns:
                        country_column_in_table = check_columns_in_table(cursor, schema, table_name, country_columns)
                    
                    # Only proceed with validation if country_column_in_table exists
                    if country_column_in_table:
                        # Load data using pandas for better region code analysis
                        cursor.execute(f"""
                            SELECT "{country_column_in_table}", "{column_name}"
                            FROM "{schema}"."{table_name}"
                        """)
                        # Fetch all data and create DataFrame
                        data = cursor.fetchall()
                        df = pd.DataFrame(data, columns=[country_column_in_table, column_name])

                        # Basic statistics - null count
                        null_count = df[column_name].isnull().sum()
                        
                        # Validate region codes
                        non_null_data = df[column_name].dropna()
                        invalid_region_codes = []
                        invalid_region_code_count = 0
                        mismatched_region_codes = []
                        mismatched_region_code_count = 0

                        if len(non_null_data) > 0:
                            # Get all valid subdivision codes from pycountry
                            valid_subdivisions = set()
                            for subdivision in pycountry.subdivisions:
                                if hasattr(subdivision, 'code'):
                                    valid_subdivisions.add(subdivision.code.upper().split('-')[-1])

                            # Convert to string and strip whitespace, then to uppercase
                            non_null_data_str = non_null_data.astype(str).str.strip().str.upper()
                            
                            # Check validity of region codes
                            not_in_valid_mask = ~non_null_data_str.isin(valid_subdivisions)
                            
                            # Extract unique invalid region codes (original values as fetched from DB)
                            invalid_region_codes = non_null_data[not_in_valid_mask].unique().tolist()
                            invalid_region_code_count = len(invalid_region_codes)
                            
                            # Check for region-country code mismatches
                            try:
                                # Get data for both region and country columns where both are not null
                                cursor.execute(f"""
                                    SELECT "{country_column_in_table}", "{column_name}"
                                    FROM "{schema}"."{table_name}"
                                    WHERE "{column_name}" IS NOT NULL AND "{country_column_in_table}" IS NOT NULL
                                """)
                                region_country_data = cursor.fetchall()
                                
                                if region_country_data:
                                    df_region_country = pd.DataFrame(region_country_data, columns=[country_column_in_table, column_name])
                                    
                                    # Create mapping from country code to valid region codes
                                    country_to_regions = {}
                                    for subdivision in pycountry.subdivisions:
                                        if hasattr(subdivision, 'country_code') and hasattr(subdivision, 'code'):
                                            country_code = subdivision.country_code.upper()
                                            region_code = subdivision.code.upper().split('-')[-1]
                                            if country_code not in country_to_regions:
                                                country_to_regions[country_code] = set()
                                            country_to_regions[country_code].add(region_code)
                                    
                                    # Process data for mismatch checking
                                    non_null_region_country = df_region_country.dropna()
                                    
                                    if len(non_null_region_country) > 0:
                                        # Convert country values to alpha_2 codes using bulk conversion
                                        country_values = non_null_region_country[country_column_in_table].astype(str).str.strip()
                                        country_values_list = country_values.tolist()
                                        # Bulk convert using country_converter
                                        country_codes_alpha2 = cc.convert(country_values_list, to='ISO2', not_found=None)
                                        country_codes_alpha2 = pd.Series(country_codes_alpha2)
                                        
                                        # Convert to uppercase for comparison
                                        country_codes_upper = country_codes_alpha2.astype(str).str.upper()
                                        region_codes_upper = non_null_region_country[column_name].astype(str).str.strip().str.upper()
                                        
                                        # Find mismatches where region code doesn't belong to the country
                                        mismatch_records = []
                                        for idx, row in non_null_region_country.iterrows():
                                            country_code = country_codes_upper.iloc[idx]
                                            region_code = region_codes_upper.iloc[idx]
                                            
                                            # Skip if country conversion failed (None or 'NONE')
                                            if country_code and country_code != 'NONE' and country_code in country_to_regions:
                                                if region_code not in country_to_regions[country_code]:
                                                    mismatch_records.append({
                                                        column_name: row[column_name],
                                                        country_column_in_table: row[country_column_in_table]
                                                    })
                                        
                                        # Get unique mismatched combinations
                                        if mismatch_records:
                                            df_mismatches = pd.DataFrame(mismatch_records)
                                            mismatched_region_codes = (
                                                df_mismatches
                                                .drop_duplicates()
                                                .to_dict('records')
                                            )
                                            mismatched_region_code_count = len(mismatched_region_codes)
                                        
                            except Exception as e:
                                # If there's an error in mismatch checking, set defaults
                                mismatched_region_codes = []
                                mismatched_region_code_count = 0
                            
                            # Get distinct regions grouped by country
                            regions_counts = df.groupby([country_column_in_table, column_name]).size()
                            distinct_regions = [
                                {
                                    "value": {
                                        country_column_in_table: str(country_val),
                                        column_name: str(region_val)
                                    },
                                    "count": int(count)
                                }
                                for (country_val, region_val), count in regions_counts.items()
                            ]

                        return Response({
                            "column_name": column_name,
                            "column_category": column_category,
                            "column_in_table": country_column_in_table,
                            "statistics": {
                                "null_count": int(null_count),
                                "total_count": len(df),
                                "distinct_regions": distinct_regions,
                                "invalid_region_category_count": invalid_region_code_count,
                                "invalid_region_categories": invalid_region_codes,
                                "mismatched_region_category_count": int(mismatched_region_code_count),
                                "mismatched_region_categories": mismatched_region_codes
                            }
                        })
                    else:
                        # If country_column_in_table doesn't exist, return response indicating it's required
                        return Response({
                            "column_name": column_name,
                            "column_category": column_category,
                            "column_in_table": None,
                            "error": "country_column_in_table is required for region_category validation"
                        }, status=400)

                elif column_category == 'postal_code':
                    # First check if country_code column exists in the table
                    country_columns = get_category_columns('country_code')
                    country_column_in_table = None
                    if country_columns:
                        country_column_in_table = check_columns_in_table(cursor, schema, table_name, country_columns)
                    
                    # Check if country column exists
                    if country_column_in_table:
                        # Load both country_code and postal_code columns together
                        print(country_column_in_table, column_name)
                        cursor.execute(f"""
                            SELECT "{country_column_in_table}", "{column_name}"
                            FROM "{schema}"."{table_name}"
                        """)
                        data = cursor.fetchall()
                        df = pd.DataFrame(data, columns=[country_column_in_table, column_name])

                        # # Check if maximum character length in postal_code column is more than three
                        # max_char_len = df[country_column_in_table].dropna().astype(str).apply(len).max() if not df.empty else 0

                        # if max_char_len > 3:
                        #     return Response(
                        #         {"error": "Maximum character length in this column is less than three."},
                        #         status=status.HTTP_400_BAD_REQUEST
                        #     )
                        
                        # Basic statistics
                        null_count = df[column_name].isnull().sum()
                        non_null_count = df[column_name].notna().sum()
                        total_count = len(df)

                        df_postal_country = df.dropna(subset=[country_column_in_table, column_name]).copy()
                        if len(df_postal_country) > 0:
                            postal_counts = df_postal_country.groupby([country_column_in_table, column_name]).size()
                            distinct_postal_codes = [
                                {
                                    "value": {
                                        country_column_in_table: str(country_val),
                                        column_name: str(postal_val)
                                    },
                                    "count": int(count)
                                }
                                for (country_val, postal_val), count in postal_counts.items()
                            ]
            
                        
                        # Get non-null data for validation - both postal_code and country must be present
                        df_validation = df.dropna(subset=[column_name, country_column_in_table]).copy()
                        
                        if len(df_validation) > 0:
                            # Cache for Nominatim objects
                            nomi_cache = {}

                            def get_alpha2_country(country):
                                """Convert country name or code to ISO alpha-2 code."""
                                if not country:
                                    return None
                                country = country.strip()
                                # If already 2 letters, assume it's alpha-2
                                if len(country) == 2:
                                    return country.upper()
                                try:
                                    return pycountry.countries.lookup(country).alpha_2
                                except LookupError:
                                    return None

                            def is_valid_postal(country_input, postal_code):
                                try:
                                    if postal_code is None or str(postal_code).strip() == "":
                                        return False
                                    
                                    country_code = get_alpha2_country(country_input)
                                    if not country_code:
                                        return False
                                    
                                    # Use cached Nominatim object
                                    if country_code not in nomi_cache:
                                        nomi_cache[country_code] = pgeocode.Nominatim(country_code)
                                    
                                    nomi = nomi_cache[country_code]
                                    loc = nomi.query_postal_code(str(postal_code))
                                    return not pd.isna(loc.get('country_code'))
                                except:
                                    return False
                            df_validation['is_valid'] = df_validation.apply(lambda row: is_valid_postal(row[country_column_in_table], row[column_name]), axis=1)

                            # Step 5: Extract all mismatched records (invalid postal codes)
                            df_mismatched = df_validation[~df_validation['is_valid']].copy()
                            
                            if len(df_mismatched) > 0:
                                # Get unique mismatched postal codes with their country codes
                                df_mismatched_unique = df_mismatched[[country_column_in_table, column_name]].drop_duplicates()
                                mismatched_postal_codes = df_mismatched_unique.to_dict('records')
                                mismatched_postal_code_count = len(mismatched_postal_codes)
                            else:
                                mismatched_postal_codes = []
                                mismatched_postal_code_count = 0
                        else:
                            # No non-null data to validate
                            mismatched_postal_codes = []
                            mismatched_postal_code_count = 0

                    
                    else:
                        # No country_code column exists - cannot validate without country
                        return Response({
                            "error": "country_code column required for postal_code validation"
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    return Response({
                        "column_name": column_name,
                        "column_category": column_category,
                        "column_in_table": country_column_in_table,
                        "statistics": {
                            "null_count": int(null_count),
                            "non_null_count": int(non_null_count),
                            "total_count": int(total_count),
                            "distinct_postal_codes": distinct_postal_codes,
                            "mismatched_postal_code_count": int(mismatched_postal_code_count),
                            "mismatched_postal_codes": mismatched_postal_codes
                        }
                    })

                elif column_category == 'units':
                    # Load data using pandas for better unit analysis
                    cursor.execute(f"""
                        SELECT "{column_name}"
                        FROM "{schema}"."{table_name}"
                    """)
                    
                    # Fetch all data and create DataFrame
                    data = cursor.fetchall()
                    df = pd.DataFrame(data, columns=[column_name])
                    
                    # Basic statistics
                    null_count = df[column_name].isnull().sum()
                    
                    # Get distinct units and their counts
                    unit_counts = df[column_name].dropna().astype(str).str.strip().value_counts()
                    distinct_values = [
                        {"value": str(unit), "count": int(count)}
                        for unit, count in unit_counts.items()
                    ]
                    
                    # Validate units using pint
                    invalid_units = []
                    
                    non_null_data = df[column_name].dropna()
                    if len(non_null_data) > 0:
                        # Create UnitRegistry instance
                        ureg = pint.UnitRegistry()
                        
                        # Convert to string and strip whitespace
                        non_null_data_str = non_null_data.astype(str).str.strip()
                        
                        # Function to validate a unit string (case-insensitive)
                        def is_valid_unit(unit_str):
                            """Validate if a unit string is a valid pint unit expression"""
                            try:
                                # Try to parse the expression in lowercase for validation
                                ureg.parse_expression(unit_str.lower())
                                return True
                            except Exception:
                                return False
                        
                        # Apply validation to each unit
                        is_valid_mask = non_null_data_str.apply(is_valid_unit)
                        
                        # Get invalid units (preserve original case)
                        invalid_mask = ~is_valid_mask
                        invalid_units = non_null_data_str[invalid_mask].unique().tolist()
                    
                    return Response({
                        "column_name": column_name,
                        "column_category": column_category,
                        "statistics": {
                            "null_count": int(null_count),
                            "distinct_values": distinct_values,
                            "invalid_units": invalid_units
                        }
                    })
                # else:
                #     cursor.execute(f"""
                #         SELECT "{column_name}"
                #         FROM "{schema}"."{table_name}"
                #     """)
                    
                #     # Fetch all data and create DataFrame
                #     data = cursor.fetchall()
                #     df = pd.DataFrame(data, columns=[column_name])
                    
                #     # Basic statistics
                #     null_count = df[column_name].isnull().sum() 
                #     category = get_column_info(column_name)   

        except Exception as e:
            return Response({
                "error": f"error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ColumnSequenceListView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            table_name = request.query_params.get('table_name')
            user = request.user
            customer = user.cust_id
            if not customer:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not table_name:
                return Response(
                    {"error": "table_name is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            with customer_connection.cursor() as cursor:
                sql = """
                    SELECT seq_name, username, scope
                    FROM "GENERAL"."tbl_col_seq"
                    WHERE table_name = %s
                    AND (
                        username = %s
                        OR scope = 'G'
                    )
                    ORDER BY seq_name
                """
                cursor.execute(sql, (table_name, user.email))
                rows = cursor.fetchall()
                result = [
                    {
                        "seq_name": row[0],
                        "username": row[1],
                        "scope": row[2],
                        "is_owner": row[1] == user.email or row[2] == 'G'  # Global sequences can be edited by anyone
                    }
                    for row in rows
                ]
            customer_connection.close()
            return Response({"sequences": result})
        except Exception as e:
            return Response({
                "error": f"error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
class ColumnSequenceView(APIView):
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            table_name = request.query_params.get('table_name')
            user = request.user
            customer = user.cust_id
            if not customer:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not table_name:
                return Response(
                    {"error": "table_name and schema are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            with customer_connection.cursor() as cursor:
                # Fetch user-specific sequences and global ones for table
                sql = f"""
                    SELECT seq_name, sequence
                    FROM "GENERAL"."tbl_col_seq"
                    WHERE table_name = %s
                    AND (
                        username = %s
                        OR scope = 'G'
                    )
                    ORDER BY 
                        seq_name
                """
                cursor.execute(sql, (table_name, user.email))
                rows = cursor.fetchall()
                result = [
                    {
                        "seq_name": row[0],
                        "sequence": row[1],
                    }
                    for row in rows
                ]
            customer_connection.close()
            return Response({"sequences": result})
        except Exception as e:
            return Response({
                "error": f"error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        """Create a new column sequence"""
        try:
            table_name = request.data.get('table_name')
            sequence = request.data.get('sequence')
            seq_name = request.data.get('seq_name')
            scope = request.data.get('scope', 'L')  # Default to Local scope
            user = request.user
            customer = user.cust_id
            
            if not customer:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not table_name or not sequence or not seq_name:
                return Response(
                    {"error": "table_name, sequence, and seq_name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']  
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Check if sequence name already exists for this table and user
                check_sql = """
                    SELECT seq_name FROM "GENERAL"."tbl_col_seq"
                    WHERE table_name = %s AND seq_name = %s AND username = %s
                """
                cursor.execute(check_sql, (table_name, seq_name, user.email))
                existing = cursor.fetchone()
                
                if existing:
                    customer_connection.close()
                    return Response(
                        {"error": f"Sequence name '{seq_name}' already exists for this table."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Insert new sequence
                sql = """
                    INSERT INTO "GENERAL"."tbl_col_seq" (username, table_name, sequence, seq_name, scope)
                    VALUES (%s, %s, %s, %s, %s)
                """
                cursor.execute(sql, (user.email, table_name, sequence, seq_name, scope))
            
            customer_connection.close()
            return Response({"message": f"Sequence '{seq_name}' created successfully"})
        except Exception as e:
            return Response({
                "error": f"Failed to create sequence: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def put(self, request):
        """Update an existing column sequence"""
        try:
            table_name = request.data.get('table_name')
            sequence = request.data.get('sequence')
            seq_name = request.data.get('seq_name')
            old_seq_name = request.data.get('old_seq_name')  # For renaming
            scope = request.data.get('scope', 'L')
            user = request.user
            customer = user.cust_id
            
            if not customer:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not table_name or not sequence or not seq_name:
                return Response(
                    {"error": "table_name, sequence, and seq_name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Use old_seq_name if provided (for renaming), otherwise use seq_name
                update_seq_name = old_seq_name if old_seq_name else seq_name
                
                # Check if user owns this sequence or if it's a global sequence
                check_sql = """
                    SELECT username, scope FROM "GENERAL"."tbl_col_seq"
                    WHERE table_name = %s AND seq_name = %s
                """
                cursor.execute(check_sql, (table_name, update_seq_name))
                result = cursor.fetchone()
                
                if not result:
                    customer_connection.close()
                    return Response(
                        {"error": f"Sequence '{update_seq_name}' not found."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Allow editing if user owns the sequence OR if it's a global sequence
                if result[0] != user.email and result[1] != 'G':
                    customer_connection.close()
                    return Response(
                        {"error": "You don't have permission to edit this sequence."},
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # If renaming, check if new name already exists
                if old_seq_name and old_seq_name != seq_name:
                    cursor.execute(check_sql, (table_name, seq_name))
                    if cursor.fetchone():
                        customer_connection.close()
                        return Response(
                            {"error": f"Sequence name '{seq_name}' already exists."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # Update sequence (keep original username if updating global sequence)
                update_sql = """
                    UPDATE "GENERAL"."tbl_col_seq"
                    SET sequence = %s, seq_name = %s, scope = %s
                    WHERE table_name = %s AND seq_name = %s
                """
                cursor.execute(update_sql, (sequence, seq_name, scope, table_name, update_seq_name))
            
            customer_connection.close()
            return Response({"message": f"Sequence '{seq_name}' updated successfully"})
        except Exception as e:
            return Response({
                "error": f"Failed to update sequence: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request):
        """Delete a column sequence"""
        try:
            table_name = request.data.get('table_name')
            seq_name = request.data.get('seq_name')
            user = request.user
            customer = user.cust_id
            
            if not customer:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not table_name or not seq_name:
                return Response(
                    {"error": "table_name and seq_name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True
            
            with customer_connection.cursor() as cursor:
                # Check if user owns this sequence or if it's a global sequence
                check_sql = """
                    SELECT username, scope FROM "GENERAL"."tbl_col_seq"
                    WHERE table_name = %s AND seq_name = %s
                """
                cursor.execute(check_sql, (table_name, seq_name))
                result = cursor.fetchone()
                
                if not result:
                    customer_connection.close()
                    return Response(
                        {"error": f"Sequence '{seq_name}' not found."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Allow deleting if user owns the sequence OR if it's a global sequence
                if result[0] != user.email and result[1] != 'G':
                    customer_connection.close()
                    return Response(
                        {"error": "You don't have permission to delete this sequence."},
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Delete sequence
                delete_sql = """
                    DELETE FROM "GENERAL"."tbl_col_seq"
                    WHERE table_name = %s AND seq_name = %s
                """
                cursor.execute(delete_sql, (table_name, seq_name))
            
            customer_connection.close()
            return Response({"message": f"Sequence '{seq_name}' deleted successfully"})
        except Exception as e:
            return Response({
                "error": f"Failed to delete sequence: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ValidationRulesView(APIView):
    # authentication_classes = [JWTCookieAuthentication]
    # permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            validation_rules = ValidationRules.objects.all().order_by('id')
            serializer = ValidationRulesSerializer(validation_rules, many=True)
            return Response({"validation_rules": serializer.data})
        except Exception as e:
            return Response({
                "error": f"error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        try:
            selected_ids = request.data.get('selected_ids', [])
            
            if not selected_ids:
                return Response({
                    "error": "No rules selected"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get selected rules from database
            selected_rules = ValidationRules.objects.filter(id__in=selected_ids).order_by('id')
            
            # Build regex by combining expressions
            regex_parts = ['^']
            
            for rule in selected_rules:
                if rule.expression:
                    regex_parts.append(rule.expression)
            
            # Add character class and end
            regex_parts.append('.+$')
            
            # Combine all parts
            final_regex = ''.join(regex_parts)
            
            # Print to terminal
            print("\n" + "="*80)
            print("GENERATED REGULAR EXPRESSION")
            print("="*80)
            print(f"Regex: {final_regex}")
            print("\nSelected Validation Rules:")
            for rule in selected_rules:
                print(f"  - {rule.question}: {rule.expression}")
            print("="*80 + "\n")
            
            return Response({"regex": final_regex})
            
        except Exception as e:
            print(f"Error generating regex: {str(e)}")
            return Response({
                "error": f"Failed to generate regex: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

